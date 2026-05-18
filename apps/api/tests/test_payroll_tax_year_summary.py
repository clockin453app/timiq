"""Employee tax-year pay summary XLSX export."""

from __future__ import annotations

import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.modules.auth.dependencies import require_authenticated_employee_self_service
from app.modules.auth.models import SystemRole, User
from app.modules.payroll.permissions import PayrollPermissionError
from app.modules.payroll.service import PayrollError, export_my_tax_year_pay_summary_xlsx


def _user(*, role: SystemRole = SystemRole.EMPLOYEE, company_id: uuid.UUID | None = None) -> User:
    now = datetime.now(timezone.utc)
    return User(
        id=uuid.uuid4(),
        company_id=company_id or uuid.uuid4(),
        email="employee@example.com",
        password_hash="hashed",
        system_role=role,
        is_active=True,
        created_at=now,
        updated_at=now,
    )


def _paid_item(*, user_id: uuid.UUID, company_id: uuid.UUID) -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid.uuid4(),
        user_id=user_id,
        company_id=company_id,
        status="paid",
        paid_at=datetime(2025, 5, 30, 9, 0, tzinfo=timezone.utc),
        gross_amount=Decimal("1000.00"),
        tax_amount=Decimal("200.00"),
        display_tax_amount=None,
        net_amount=Decimal("750.00"),
        display_net_amount=None,
        other_deductions_amount=Decimal("50.00"),
        payment_mode="net_payment",
        tax_rate_snapshot=Decimal("20.00"),
    )


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def test_tax_year_summary_requires_employee_role() -> None:
    with pytest.raises(PayrollPermissionError):
        export_my_tax_year_pay_summary_xlsx(
            MagicMock(),
            _user(role=SystemRole.ADMIN),
            tax_year="2025-2026",
        )


def test_tax_year_summary_uses_uk_tax_year_range_and_builds_workbook() -> None:
    actor = _user()
    company_id = uuid.uuid4()
    item = _paid_item(user_id=actor.id, company_id=company_id)
    period = SimpleNamespace(week_start=date(2025, 5, 26))

    with (
        patch(
            "app.modules.payroll.service.list_paid_items_for_user_tax_year_summary",
            return_value=[(item, period)],
        ) as repo,
        patch("app.modules.payroll.service._employee_primary_name", return_value="Employee Example"),
        patch("app.modules.payroll.service._employee_tax_identifiers_for_payroll", return_value=("AB123456C", "1234567890")),
        patch("app.modules.payroll.service.get_company_by_id", return_value=SimpleNamespace(name="Acme Ltd")),
        patch("app.modules.payroll.service.create_internal_audit_event"),
    ):
        body = export_my_tax_year_pay_summary_xlsx(MagicMock(), actor, tax_year="2025-2026")

    kwargs = repo.call_args.kwargs
    assert kwargs["user_id"] == actor.id
    assert kwargs["paid_at_from"] == datetime(2025, 4, 6, tzinfo=timezone.utc)
    assert kwargs["paid_at_before"] == datetime(2026, 4, 6, tzinfo=timezone.utc)

    workbook = load_workbook(BytesIO(body), data_only=True)
    assert workbook.sheetnames == ["Payslips", "Companies"]
    payslips = workbook["Payslips"]
    assert payslips["D2"].value == "Employee Example"
    assert payslips["D3"].value == "AB123456C"
    assert payslips["D4"].value == "1234567890"
    assert payslips["D5"].value == "2025/2026"
    assert payslips["B12"].value == "№"
    assert payslips["G13"].value == 1000
    assert payslips["H13"].value == 200
    assert payslips["J13"].value == 250
    assert payslips["M13"].value == 750
    assert payslips["G2"].value == 1000
    assert payslips["J5"].value is None
    companies = workbook["Companies"]
    assert companies["C3"].value == "Acme Ltd"


def test_tax_year_summary_empty_year_returns_clean_workbook() -> None:
    actor = _user()
    with (
        patch("app.modules.payroll.service.list_paid_items_for_user_tax_year_summary", return_value=[]),
        patch("app.modules.payroll.service._employee_primary_name", return_value="Employee Example"),
        patch("app.modules.payroll.service._employee_tax_identifiers_for_payroll", return_value=(None, None)),
        patch("app.modules.payroll.service.create_internal_audit_event"),
    ):
        body = export_my_tax_year_pay_summary_xlsx(MagicMock(), actor, tax_year="2025-2026")

    workbook = load_workbook(BytesIO(body), data_only=True)
    payslips = workbook["Payslips"]
    assert payslips["G2"].value == 0
    assert payslips["B14"].value == "Totals"


def test_tax_year_summary_rejects_invalid_tax_year() -> None:
    with pytest.raises(PayrollError):
        export_my_tax_year_pay_summary_xlsx(
            MagicMock(),
            _user(),
            tax_year="2025-2027",
        )


@patch("app.modules.payroll.router.export_my_tax_year_pay_summary_xlsx")
def test_tax_year_summary_endpoint_response_headers(mock_export: MagicMock, client: TestClient) -> None:
    employee = _user()
    mock_export.return_value = b"PK\x03\x04 workbook"
    app.dependency_overrides[require_authenticated_employee_self_service] = lambda: employee
    try:
        response = client.get("/api/payroll/pay-history/me/tax-year-summary.xlsx?tax_year=2025-2026")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "timiq-pay-summary-2025-2026.xlsx" in response.headers.get("content-disposition", "")
        assert mock_export.call_args.kwargs["tax_year"] == "2025-2026"
    finally:
        app.dependency_overrides.clear()


def test_tax_year_summary_endpoint_requires_authentication(client: TestClient) -> None:
    response = client.get("/api/payroll/pay-history/me/tax-year-summary.xlsx?tax_year=2025-2026")
    assert response.status_code in (401, 403)
