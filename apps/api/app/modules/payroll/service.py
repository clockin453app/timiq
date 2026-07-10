"""Payroll orchestration: recalculate, approvals, exports."""

from __future__ import annotations

import csv
import html
import io
import uuid
from datetime import date, datetime, time, timedelta, timezone
from decimal import ROUND_HALF_UP, Decimal
from typing import Any
from zoneinfo import ZoneInfo

from sqlalchemy.orm import Session

from app.core.export_csv import seconds_to_hours_csv
from app.modules.audit.service import create_internal_audit_event
from app.modules.auth.models import SystemRole, User
from app.modules.auth.repository import get_user_by_id
from app.modules.accounting.repository import payroll_export_run_overlaps_date_range
from app.modules.leave import repository as leave_repo
from app.modules.companies.repository import get_company_by_id
from app.modules.companies.service import ensure_company_time_policy
from app.modules.employee_profiles.repository import get_employee_profile_by_user_id
from app.modules.onboarding.repository import (
    get_approved_onboarding_national_insurance_number,
    get_approved_onboarding_utr,
)
from app.modules.notifications.events import record_payroll_paid
from app.modules.payroll.calculation import (
    compute_money_bundle,
    normalize_payroll_payment_mode,
    policy_snapshot_dict,
    regular_overtime_seconds_payroll_week,
    resolve_effective_tax_rate_percent,
    split_regular_overtime_daily_by_work_date,
    sum_rounded_seconds_payroll_week,
    week_bounds_utc,
    work_date_in_policy,
)
from app.modules.payroll.late_shifts import (
    append_late_shift_ids_marker,
    reserved_late_shift_ids_for_user_period,
    shift_completed_after_paid_cutoff,
)
from app.modules.payroll.models import PayrollItem, PayrollPeriod
from app.modules.payroll.permissions import (
    PayrollPermissionError,
    assert_actor_can_view_payroll_item,
    assert_payroll_admin_or_administrator,
    assert_payroll_company_scope,
)
from app.modules.payroll_policies.service import effective_early_access_for_shift, effective_time_policy_for_shift
from app.modules.payroll.pdf_export import build_payroll_item_payslip_pdf, build_payroll_report_pdf
from app.modules.payroll.repository import (
    count_open_shifts_started_in_week,
    delete_pending_items_for_period,
    first_workplace_tax,
    get_item_by_id,
    get_period_by_company_week,
    invalidate_period_calculation_for_company_week,
    list_cis_employee_users_for_company,
    list_employee_users_for_company,
    list_completed_time_shifts_for_company_range,
    list_items_for_period,
    list_items_for_user_pay_history,
    list_paid_items_for_user_tax_year_summary,
    list_paid_items_for_company_payment_history,
    list_payroll_items_for_user_company_ytd_calendar_year,
    list_periods_for_company_month,
    max_employee_shift_updated_at_in_payroll_week,
    period_has_approved_item,
    period_has_paid_item,
    save_item,
    save_period,
    update_item,
)
from app.modules.time_records.calculation import compute_shift_metrics
from app.modules.time_records.repository import list_time_shifts_for_payroll_week
from app.modules.payroll.schemas import (
    PayHistoryEntry,
    PayrollItemCompanySnippet,
    PayrollItemPatchRequest,
    PayrollItemResponse,
    PayrollItemSummaryResponse,
    PayrollApprovedLeaveRow,
    PayrollLateAdjustmentRequest,
    PayrollLateShiftRow,
    PayrollLateUnpaidEmployee,
    PayrollMonthSummaryResponse,
    PayrollPaymentHistoryRow,
    PayrollPaySplit,
    PayrollPeriodSummary,
    PayrollReportAlerts,
    PayrollReportResponse,
    PayrollUndoPaidRequest,
)


class PayrollError(ValueError):
    pass


class PayrollPaidBlockingError(PayrollError):
    pass


class PayrollApprovedBlockingError(PayrollError):
    pass


class PayrollItemStateError(PayrollError):
    pass


class PayrollItemNotFoundError(Exception):
    """Missing payroll item or period (maps to HTTP 404)."""


def _decimal_or_none(value: object | None) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value))


def _effective_tax_amount_for_item(item: PayrollItem) -> Decimal | None:
    """CIS tax for summaries and payslips: zero under gross payment; net uses display override semantics."""
    mode = normalize_payroll_payment_mode(item.payment_mode)
    if mode == "gross_payment":
        if item.rate_missing:
            return None
        if _decimal_or_none(item.gross_amount) is None:
            return None
        return Decimal(0)
    display = _decimal_or_none(item.display_tax_amount)
    calculated = _decimal_or_none(item.tax_amount)
    if display is not None:
        if display == 0 and calculated is not None and calculated != 0:
            return calculated
        return display
    return calculated


def _effective_net_amount_for_item(item: PayrollItem) -> Decimal | None:
    display = _decimal_or_none(item.display_net_amount)
    if display is not None:
        return display
    return _decimal_or_none(item.net_amount)


def _payment_mode_label(payment_mode: str | None) -> str:
    mode = normalize_payroll_payment_mode(payment_mode)
    if mode == "gross_payment":
        return "Gross payment"
    return "Net payment"


def _stored_payment_mode_or_none(payment_mode: str | None) -> str | None:
    if payment_mode is None:
        return None
    raw = str(payment_mode).strip().lower()
    if raw in ("net", "net_payment", "gross", "gross_payment"):
        return normalize_payroll_payment_mode(raw)
    return None


def _stored_payment_mode_source_or_none(payment_mode_source: str | None) -> str | None:
    if payment_mode_source is None:
        return None
    raw = str(payment_mode_source).strip().lower()
    if raw in ("profile", "manual"):
        return raw
    return None


def _payment_mode_label_for_item(item: PayrollItem) -> str:
    mode = _stored_payment_mode_or_none(getattr(item, "payment_mode", None))
    if mode is None:
        return "Not provided"
    return _payment_mode_label(mode)


def _apply_payroll_item_money_after_patch(item: PayrollItem, request: PayrollItemPatchRequest) -> None:
    """Reconcile tax/net/display amounts from stored hours, snapshots, payment mode, and other deductions."""
    if item.rate_missing:
        return
    hourly = _decimal_or_none(item.hourly_rate_snapshot)
    if hourly is None:
        return
    ot_mult = _decimal_or_none(item.overtime_multiplier_snapshot) or Decimal(1)
    tax_pct = _decimal_or_none(item.tax_rate_snapshot)
    other_d = Decimal(str(item.other_deductions_amount or 0))
    mode = normalize_payroll_payment_mode(item.payment_mode)
    item.payment_mode = mode

    bundle = compute_money_bundle(
        regular_seconds=item.regular_seconds,
        overtime_seconds=item.overtime_seconds,
        hourly_rate=hourly,
        overtime_multiplier=ot_mult,
        tax_rate_percent=tax_pct,
        other_deductions=other_d,
        payment_mode=mode,
    )
    if bool(bundle["rate_missing"]):
        return

    recompute_core = request.payment_mode is not None or request.other_deductions_amount is not None

    if recompute_core or mode == "gross_payment":
        item.tax_amount = float(bundle["tax_amount"]) if bundle["tax_amount"] is not None else None
        item.net_amount = float(bundle["net_amount"]) if bundle["net_amount"] is not None else None
        item.display_tax_amount = (
            float(bundle["display_tax_amount"]) if bundle["display_tax_amount"] is not None else None
        )
        if request.display_net_amount is not None and not recompute_core:
            item.display_net_amount = float(request.display_net_amount)
            item.net_amount = float(request.display_net_amount)
        else:
            item.display_net_amount = (
                float(bundle["display_net_amount"]) if bundle["display_net_amount"] is not None else None
            )
        return

    gross = _decimal_or_none(item.gross_amount)
    if gross is None:
        return
    tax_for_net = _effective_tax_amount_for_item(item)
    if tax_for_net is None:
        return
    if request.display_net_amount is None:
        net_calc = (gross - tax_for_net - other_d).quantize(Decimal("0.01"))
        item.net_amount = float(net_calc)
        item.display_net_amount = float(net_calc)


def _employee_tax_identifiers_for_payroll(
    db_session: Session,
    user_id: uuid.UUID,
) -> tuple[str | None, str | None]:
    """NI and UTR for payslip / pay-week: profile first, then approved onboarding fallback."""
    profile = get_employee_profile_by_user_id(db_session, user_id)
    ni = (profile.national_insurance_number or "").strip() if profile is not None else None
    utr = (profile.utr_number or "").strip() if profile is not None else None
    if not ni:
        ni = get_approved_onboarding_national_insurance_number(db_session, user_id)
    if not utr:
        utr = get_approved_onboarding_utr(db_session, user_id)
    return ni or None, utr or None


def _week_end_display(week_start: date) -> date:
    return week_start + timedelta(days=6)


PARTIAL_RANGE_PAYROLL_NOTE = (
    "Payroll pay totals are weekly aggregates and are not split across partial date ranges. "
    "This export shows shift hours for the selected dates."
)


def _policy_zone_name(timezone_name: str | None) -> ZoneInfo:
    try:
        return ZoneInfo(timezone_name or "UTC")
    except Exception:
        return ZoneInfo("UTC")


def _date_range_bounds_utc(timezone_name: str, date_from: date, date_to: date) -> tuple[datetime, datetime]:
    tz = _policy_zone_name(timezone_name)
    start_local = datetime.combine(date_from, time.min, tzinfo=tz)
    end_local = datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=tz)
    return start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc)


def _payroll_week_start_for_dt(dt_utc: datetime, timezone_name: str) -> date:
    local_date = dt_utc.astimezone(_policy_zone_name(timezone_name)).date()
    return local_date - timedelta(days=local_date.weekday())


def mark_payroll_period_needs_recalculation(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
) -> bool:
    """Invalidate an existing payroll period without changing item money/status values."""
    return invalidate_period_calculation_for_company_week(
        db_session,
        company_id=company_id,
        week_start=week_start,
    )


def _complete_week_starts_in_range(date_from: date, date_to: date) -> list[date]:
    cursor = date_from + timedelta(days=(7 - date_from.weekday()) % 7)
    weeks: list[date] = []
    while cursor + timedelta(days=6) <= date_to:
        weeks.append(cursor)
        cursor += timedelta(days=7)
    return weeks


def _range_has_partial_week_portion(date_from: date, date_to: date) -> bool:
    complete_days: set[date] = set()
    for week_start in _complete_week_starts_in_range(date_from, date_to):
        for offset in range(7):
            complete_days.add(week_start + timedelta(days=offset))
    cursor = date_from
    while cursor <= date_to:
        if cursor not in complete_days:
            return True
        cursor += timedelta(days=1)
    return False


def _accounting_export_overlaps_payroll_week(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
) -> bool:
    return payroll_export_run_overlaps_date_range(
        db_session,
        company_id=company_id,
        range_start=week_start,
        range_end=_week_end_display(week_start),
    )


def _payroll_approved_leave_rows(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
    filter_user_id: uuid.UUID | None,
) -> list[PayrollApprovedLeaveRow]:
    w_end = week_start + timedelta(days=6)
    rows = leave_repo.list_leave_overlapping_week(
        db_session,
        company_id=company_id,
        week_start=week_start,
        week_end=w_end,
        statuses=("approved",),
        user_id=filter_user_id,
    )
    out: list[PayrollApprovedLeaveRow] = []
    for r in rows:
        email, name, _jt = _employee_display(db_session, r.user_id)
        out.append(
            PayrollApprovedLeaveRow(
                user_id=r.user_id,
                employee_email=email,
                employee_name=name,
                leave_type=r.leave_type,
                date_from=r.date_from,
                date_to=r.date_to,
                total_days=Decimal(str(r.total_days)),
            )
        )
    return out


def _utc_dt_display_for_payslip(dt: datetime | None) -> str:
    if dt is None:
        return ""
    aware = dt if dt.tzinfo is not None else dt.replace(tzinfo=timezone.utc)
    return html.escape(aware.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))


def _employee_primary_name(db_session: Session, user_id: uuid.UUID) -> str:
    email, name, _jt = _employee_display(db_session, user_id)
    if name and str(name).strip():
        return str(name).strip()
    return email or "Employee"


def _compute_ytd_for_item(
    db_session: Session,
    item: PayrollItem,
    period: PayrollPeriod,
) -> tuple[Decimal, Decimal]:
    calendar_year = period.week_start.year
    rows = list_payroll_items_for_user_company_ytd_calendar_year(
        db_session,
        user_id=item.user_id,
        company_id=item.company_id,
        calendar_year=calendar_year,
        through_week_start=period.week_start,
    )
    gross_sum = Decimal(0)
    cis_sum = Decimal(0)
    for row in rows:
        if row.gross_amount is not None:
            gross_sum += Decimal(str(row.gross_amount))
        t = _effective_tax_amount_for_item(row)
        if t is not None:
            cis_sum += t
    return gross_sum, cis_sum


def _load_item_period_owner(
    db_session: Session,
    item_id: uuid.UUID,
) -> tuple[PayrollItem, PayrollPeriod, User]:
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollItemNotFoundError("Payroll item not found.")
    period = db_session.get(PayrollPeriod, item.period_id)
    if period is None:
        raise PayrollItemNotFoundError("Payroll item not found.")
    owner = get_user_by_id(db_session, item.user_id)
    if owner is None:
        raise PayrollItemNotFoundError("Payroll item not found.")
    return item, period, owner


def get_payroll_item_summary(db_session: Session, actor: User, item_id: uuid.UUID) -> PayrollItemSummaryResponse:
    item, period, owner = _load_item_period_owner(db_session, item_id)
    assert_actor_can_view_payroll_item(actor, item, owner)
    ytd_pay, ytd_cis = _compute_ytd_for_item(db_session, item, period)
    company = get_company_by_id(db_session, item.company_id)
    cis = _effective_tax_amount_for_item(item)
    net_eff = _effective_net_amount_for_item(item)
    owner_email = owner.email
    ni_val, utr_val = _employee_tax_identifiers_for_payroll(db_session, item.user_id)
    return PayrollItemSummaryResponse(
        item_id=item.id,
        company=PayrollItemCompanySnippet(
            id=item.company_id,
            name=company.name if company is not None else "Company",
        ),
        employee_display_name=_employee_primary_name(db_session, item.user_id),
        employee_email=owner_email,
        timezone_name=period.timezone_name,
        week_start=period.week_start,
        week_end=_week_end_display(period.week_start),
        status=item.status,
        approved_at=item.approved_at,
        paid_at=item.paid_at,
        payment_mode=_stored_payment_mode_or_none(item.payment_mode),
        payment_mode_label=_payment_mode_label_for_item(item),
        regular_seconds=item.regular_seconds,
        overtime_seconds=item.overtime_seconds,
        rounded_total_seconds=item.rounded_total_seconds,
        gross_amount=_decimal_or_none(item.gross_amount),
        cis_tax_amount=cis,
        net_amount=net_eff,
        other_deductions_amount=Decimal(str(item.other_deductions_amount or 0)),
        hourly_rate_snapshot=_decimal_or_none(item.hourly_rate_snapshot),
        rate_missing=item.rate_missing,
        ytd_taxable_pay=ytd_pay,
        ytd_cis_deducted=ytd_cis,
        can_open_payslip=True,
        national_insurance_number=ni_val,
        utr_number=utr_val,
    )


def render_payroll_item_payslip_html(db_session: Session, actor: User, item_id: uuid.UUID) -> str:
    item, period, owner = _load_item_period_owner(db_session, item_id)
    assert_actor_can_view_payroll_item(actor, item, owner)
    ytd_pay, ytd_cis = _compute_ytd_for_item(db_session, item, period)
    company = get_company_by_id(db_session, item.company_id)
    cname = html.escape(company.name if company is not None else "Company")
    ename = html.escape(_employee_primary_name(db_session, item.user_id))
    ni, utr = _employee_tax_identifiers_for_payroll(db_session, item.user_id)
    ni_display = html.escape(ni) if ni else "Not provided"
    utr_display = html.escape(utr) if utr else "Not provided"
    cis = _effective_tax_amount_for_item(item)
    net_eff = _effective_net_amount_for_item(item)
    gross = _decimal_or_none(item.gross_amount)
    tot_h = item.rounded_total_seconds / 3600
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    mode_label = html.escape(_payment_mode_label_for_item(item))
    if item.paid_at is not None:
        status_label = "Paid"
        pay_date_label = item.paid_at.strftime("%Y-%m-%d")
    elif item.approved_at is not None:
        status_label = "Approved"
        pay_date_label = item.approved_at.strftime("%Y-%m-%d")
    else:
        status_label = "Not provided"
        pay_date_label = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    week_end = _week_end_display(period.week_start)
    week_number = period.week_start.isocalendar().week
    week_label = (
        f"Week {week_number} "
        f"({period.week_start.strftime('%d %b')} \u2013 {week_end.strftime('%d %b %Y')})"
    )

    net_display = f"£{net_eff:.2f}" if net_eff is not None else "—"
    gross_display = f"£{gross:.2f}" if gross is not None else "—"
    cis_display = f"£{cis:.2f}" if cis is not None else "—"
    statement_heading = "CIS Pay Statement" if cis is not None and cis != 0 else "Payslip"

    html_out = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Payslip — {cname}</title>
<style>
html {{ box-sizing: border-box; }}
*, *::before, *::after {{ box-sizing: inherit; }}
body {{
  margin: 0;
  padding: 0;
  background: #f5f7fb;
  color: #111827;
  font-family: system-ui, -apple-system, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  font-size: clamp(14px, 2.8vw, 16px);
  line-height: 1.5;
  -webkit-text-size-adjust: 100%;
}}
.payslip-wrap {{
  width: 100%;
  max-width: 100%;
  margin: 0 auto;
  padding: 10px 16px 32px;
}}
@media (min-width: 768px) {{
  .payslip-wrap {{ padding: 12px 32px 44px; }}
}}
.action-row {{
  width: min(100%, 980px);
  margin: 0 auto;
  display: grid;
  grid-template-columns: minmax(9rem, auto) minmax(0, 1fr) 72px;
  align-items: center;
  gap: 10px;
}}
.back-btn {{
  border: 0;
  background: transparent;
  color: #475569;
  cursor: pointer;
  font: inherit;
  font-weight: 700;
  padding: 0.65rem 0;
  text-align: left;
}}
.print-btn {{
  border: 0;
  border-radius: 10px;
  background: #1464ee;
  color: #fff;
  cursor: pointer;
  font: inherit;
  font-weight: 800;
  padding: 0.78rem 1rem;
  text-align: center;
  box-shadow: 0 5px 14px rgba(20, 100, 238, 0.2);
}}
.payslip-card {{
  width: min(100%, 980px);
  margin: 18px auto 0;
  background: #fff;
  border: 1px solid #d9e0ea;
  padding: 24px 24px 26px;
  box-shadow: 0 16px 34px rgba(15, 23, 42, 0.08);
}}
@media (min-width: 768px) {{
  .payslip-card {{ padding: 30px 34px 28px; }}
}}
.payslip-head {{
  display: grid;
  grid-template-columns: minmax(0, 1.1fr) minmax(250px, 0.9fr);
  gap: 24px;
  padding-bottom: 20px;
  margin-bottom: 18px;
  border-bottom: 1px solid #e5e7eb;
}}
.co-name {{
  font-size: clamp(1.35rem, 3vw, 1.65rem);
  font-weight: 850;
  letter-spacing: 0.01em;
  text-transform: uppercase;
  word-break: break-word;
}}
.eyebrow {{
  color: #64748b;
  font-size: 0.83rem;
  font-weight: 600;
  margin-top: 0.15rem;
}}
.employee-name {{
  font-size: 1.15rem;
  font-weight: 800;
  margin-top: 1rem;
}}
.identity-line {{
  color: #64748b;
  font-size: 0.9rem;
  font-weight: 650;
  margin-top: 0.18rem;
}}
.payslip-head-right {{ text-align: right; }}
.doc-type {{
  color: #111827;
  font-size: clamp(1.35rem, 3vw, 1.65rem);
  font-weight: 850;
  line-height: 1.15;
}}
.week-title {{
  color: #334155;
  font-size: 1.02rem;
  font-weight: 800;
  margin-top: 0.35rem;
}}
.generated {{
  color: #64748b;
  font-size: 0.78rem;
  font-weight: 650;
  margin-top: 0.75rem;
}}
.statement-body {{
  display: grid;
  grid-template-columns: minmax(0, 1fr) minmax(280px, 0.92fr);
  gap: 56px;
  padding: 0.9rem 0 0.7rem;
}}
.section-title {{
  color: #356e9f;
  font-size: 0.78rem;
  font-weight: 850;
  letter-spacing: 0.08em;
  margin: 0 0 0.65rem;
  text-transform: uppercase;
}}
.row {{
  display: grid;
  grid-template-columns: minmax(130px, 1fr) minmax(120px, auto);
  gap: 18px;
  padding: 0.31rem 0;
}}
.row-label {{
  color: #475569;
  font-weight: 650;
}}
.row-value {{
  color: #111827;
  font-weight: 750;
  text-align: right;
  font-variant-numeric: tabular-nums;
}}
.total-row {{
  border-top: 1px solid #e5e7eb;
  margin-top: 0.55rem;
  padding-top: 0.72rem;
}}
.total-row .row-label,
.total-row .row-value {{
  font-weight: 850;
}}
.pay-date {{
  color: #334155;
  font-size: 0.88rem;
  font-weight: 850;
  letter-spacing: 0.02em;
  margin: 0 0 0.9rem;
  text-transform: uppercase;
}}
.accent-line {{
  height: 14px;
  margin-top: 1rem;
  background: #2f6f9e;
}}
@media (max-width: 760px) {{
  .action-row {{ grid-template-columns: 1fr; }}
  .print-btn {{ width: 100%; }}
  .payslip-card {{ margin-top: 12px; padding: 20px 16px 22px; }}
  .payslip-head,
  .statement-body {{ grid-template-columns: 1fr; gap: 20px; }}
  .payslip-head-right {{ text-align: left; }}
}}
@media print {{
  body {{ background: #fff; font-size: 11pt; }}
  .payslip-wrap {{ max-width: none; padding: 0; }}
  .action-row {{ display: none; }}
  .payslip-card {{ width: 100%; border: none; box-shadow: none; padding: 0; }}
  @page {{ size: A4; margin: 12mm; }}
}}
</style>
<script>
function closeOrReturnToPayroll() {{
  function fallbackNavigate() {{
    if (!window.closed) {{
      if (window.history.length > 1) {{
        window.history.back();
      }} else {{
        window.location.href = "/payroll-report";
      }}
    }}
  }}

  if (window.opener && !window.opener.closed) {{
    window.close();
    setTimeout(fallbackNavigate, 150);
    return;
  }}

  if (window.history.length > 1) {{
    window.history.back();
    return;
  }}

  window.close();
  setTimeout(fallbackNavigate, 150);
}}
</script></head><body>
<div class="payslip-wrap">
  <div class="action-row">
    <button class="back-btn" onclick="closeOrReturnToPayroll()" type="button">\u2190 Close payslip</button>
    <button class="print-btn" onclick="window.print()" type="button">Save / Print Payslip</button>
    <span aria-hidden="true"></span>
  </div>
  <div class="payslip-card">
    <header class="payslip-head">
      <div class="payslip-head-left">
        <div class="co-name">{cname}</div>
        <div class="eyebrow">Company</div>
        <div class="employee-name">{ename}</div>
        <div class="identity-line">UTR: {utr_display}</div>
        <div class="identity-line">National Insurance: {ni_display}</div>
      </div>
      <div class="payslip-head-right">
        <div class="doc-type">{html.escape(statement_heading)}</div>
        <div class="week-title">{html.escape(week_label)}</div>
        <div class="generated">Generated: {html.escape(generated)}</div>
      </div>
    </header>
    <main class="statement-body">
      <section>
        <h2 class="section-title">Pay Summary</h2>
        <div class="row"><span class="row-label">Status</span><span class="row-value">{html.escape(status_label)}</span></div>
        <div class="row"><span class="row-label">Payment type</span><span class="row-value">{mode_label}</span></div>
        <div class="row"><span class="row-label">Hours worked</span><span class="row-value">{tot_h:.2f}</span></div>
        <div class="row"><span class="row-label">Gross pay</span><span class="row-value">{gross_display}</span></div>
        <div class="row"><span class="row-label">CIS tax</span><span class="row-value">{cis_display}</span></div>
        <div class="row total-row"><span class="row-label">Total net pay</span><span class="row-value">{net_display}</span></div>
      </section>
      <section>
        <p class="pay-date">Pay date: {html.escape(pay_date_label)}</p>
        <h2 class="section-title">Year to Date</h2>
        <div class="row"><span class="row-label">Taxable Pay</span><span class="row-value">£{ytd_pay:.2f}</span></div>
        <div class="row"><span class="row-label">CIS deducted YTD</span><span class="row-value">£{ytd_cis:.2f}</span></div>
      </section>
    </main>
    <div class="accent-line"></div>
  </div>
</div>
</body></html>"""

    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.payslip_viewed",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={
            "item_id": str(item.id),
            "owner_user_id": str(item.user_id),
            "company_id": str(item.company_id),
            "actor_user_id": str(actor.id),
            "as_admin": actor.id != item.user_id,
        },
    )
    return html_out


def render_payroll_item_payslip_pdf(db_session: Session, actor: User, item_id: uuid.UUID) -> tuple[bytes, date]:
    item, period, owner = _load_item_period_owner(db_session, item_id)
    assert_actor_can_view_payroll_item(actor, item, owner)
    ytd_pay, ytd_cis = _compute_ytd_for_item(db_session, item, period)
    company = get_company_by_id(db_session, item.company_id)
    ni, utr = _employee_tax_identifiers_for_payroll(db_session, item.user_id)
    if item.paid_at is not None:
        status_label = "Paid"
        pay_date_label = item.paid_at.strftime("%Y-%m-%d")
    elif item.approved_at is not None:
        status_label = "Approved"
        pay_date_label = item.approved_at.strftime("%Y-%m-%d")
    else:
        status_label = "Not provided"
        pay_date_label = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    week_end = _week_end_display(period.week_start)
    week_label = (
        f"Week {period.week_start.isocalendar().week} "
        f"({period.week_start.strftime('%d %b')} \u2013 {week_end.strftime('%d %b %Y')})"
    )

    body = build_payroll_item_payslip_pdf(
        company_name=company.name if company is not None else "Company",
        employee_name=_employee_primary_name(db_session, item.user_id),
        employee_email=owner.email,
        national_insurance_number=ni,
        utr_number=utr,
        week_start=period.week_start,
        week_end=week_end,
        timezone_name=period.timezone_name,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        status_label=status_label,
        pay_date_label=pay_date_label,
        week_label=week_label,
        payment_mode_label=_payment_mode_label_for_item(item),
        regular_hours=item.regular_seconds / 3600,
        overtime_hours=item.overtime_seconds / 3600,
        total_hours=item.rounded_total_seconds / 3600,
        gross_amount=_decimal_or_none(item.gross_amount),
        cis_tax_amount=_effective_tax_amount_for_item(item),
        other_deductions_amount=Decimal(str(item.other_deductions_amount or 0)),
        additions_amount=Decimal("0.00"),
        net_amount=_effective_net_amount_for_item(item),
        ytd_taxable_pay=ytd_pay,
        ytd_cis_deducted=ytd_cis,
    )
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.payslip_pdf_downloaded",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={
            "item_id": str(item.id),
            "owner_user_id": str(item.user_id),
            "company_id": str(item.company_id),
            "actor_user_id": str(actor.id),
            "as_admin": actor.id != item.user_id,
        },
    )
    return body, period.week_start


def _employee_display(
    db_session: Session,
    user_id: uuid.UUID,
) -> tuple[str | None, str | None, str | None]:
    user = get_user_by_id(db_session, user_id)
    if user is None:
        return None, None, None
    profile = get_employee_profile_by_user_id(db_session, user_id)
    if profile is None:
        return user.email, None, None
    first = (profile.first_name or "").strip()
    last = (profile.last_name or "").strip()
    name = f"{first} {last}".strip() if first or last else None
    job_title = (profile.job_title or "").strip() or None
    return user.email, name, job_title


def _item_regular_overtime_pay_components(item: PayrollItem) -> tuple[Decimal, Decimal]:
    if item.rate_missing:
        return Decimal(0), Decimal(0)
    hourly = _decimal_or_none(item.hourly_rate_snapshot)
    if hourly is None:
        return Decimal(0), Decimal(0)
    mult = _decimal_or_none(item.overtime_multiplier_snapshot) or Decimal(1)
    reg_h = Decimal(item.regular_seconds) / Decimal(3600)
    ot_h = Decimal(item.overtime_seconds) / Decimal(3600)
    reg_p = (reg_h * hourly).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    ot_p = (ot_h * hourly * mult).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    return reg_p, ot_p


def _build_pay_split(items: list[PayrollItem]) -> PayrollPaySplit:
    reg = Decimal(0)
    ot = Decimal(0)
    for i in items:
        rp, op = _item_regular_overtime_pay_components(i)
        reg += rp
        ot += op
    tg = Decimal(0)
    has_gross = False
    for i in items:
        if i.gross_amount is not None:
            tg += Decimal(str(i.gross_amount))
            has_gross = True
    return PayrollPaySplit(
        regular_pay=reg,
        overtime_pay=ot,
        other_pay=Decimal(0),
        total_gross=tg if has_gross else None,
    )


def _item_missing_hourly_rate_for_approval(item: PayrollItem) -> bool:
    if bool(getattr(item, "rate_missing", False)):
        return True
    if not hasattr(item, "hourly_rate_snapshot"):
        return False
    hourly = _decimal_or_none(getattr(item, "hourly_rate_snapshot"))
    return hourly is None or hourly <= 0


def _item_missing_required_payroll_setup(item: PayrollItem) -> bool:
    if not hasattr(item, "tax_rate_snapshot"):
        return False
    if normalize_payroll_payment_mode(getattr(item, "payment_mode", None)) != "net_payment":
        return False
    return _decimal_or_none(getattr(item, "tax_rate_snapshot")) is None


def _item_has_actionable_pending_payroll_value(item: PayrollItem) -> bool:
    if getattr(item, "status", None) != "pending":
        return False
    if int(getattr(item, "rounded_total_seconds", 0) or 0) > 0:
        return True
    if int(getattr(item, "regular_seconds", 0) or 0) > 0:
        return True
    if int(getattr(item, "overtime_seconds", 0) or 0) > 0:
        return True
    money_values = (
        _decimal_or_none(getattr(item, "gross_amount", None)),
        _effective_tax_amount_for_item(item),
        _effective_net_amount_for_item(item),
        _decimal_or_none(getattr(item, "other_deductions_amount", None)),
    )
    return any(value is not None and value > 0 for value in money_values)


def _missing_profile_hourly_rate_count(db_session: Session, all_items: list[PayrollItem]) -> int:
    missing = 0
    seen_user_ids: set[uuid.UUID] = set()
    for item in all_items:
        user_id = getattr(item, "user_id", None)
        if user_id is None or user_id in seen_user_ids:
            continue
        seen_user_ids.add(user_id)
        profile = get_employee_profile_by_user_id(db_session, user_id)
        hourly = _decimal_or_none(getattr(profile, "hourly_rate", None)) if profile is not None else None
        if hourly is None or hourly <= 0:
            missing += 1
    return missing


def _missing_tax_identifier_counts(db_session: Session, all_items: list[PayrollItem]) -> tuple[int, int]:
    utr_missing = 0
    nino_missing = 0
    seen_user_ids: set[uuid.UUID] = set()
    for item in all_items:
        user_id = getattr(item, "user_id", None)
        if user_id is None or user_id in seen_user_ids:
            continue
        seen_user_ids.add(user_id)
        ni, utr = _employee_tax_identifiers_for_payroll(db_session, user_id)
        if not utr:
            utr_missing += 1
        if not ni:
            nino_missing += 1
    return utr_missing, nino_missing


_POLICY_SNAPSHOT_COMPARE_KEYS = (
    "standard_start_time",
    "overtime_after_hours",
    "overtime_multiplier",
    "rounding_increment_minutes",
    "rounding_mode",
    "break_deduction_minutes",
    "break_deduction_after_minutes",
    "rule_effective_from",
    "timezone_name",
)


def _decimal_snapshots_equal(current: Decimal | None, stored: float | Decimal | None) -> bool:
    current_q = _decimal_or_none(current)
    stored_q = _decimal_or_none(stored)
    if current_q is None and stored_q is None:
        return True
    if current_q is None or stored_q is None:
        return False
    return current_q.quantize(Decimal("0.0001")) == stored_q.quantize(Decimal("0.0001"))


def _normalize_policy_snapshot_field(key: str, value: object | None) -> str:
    if value is None:
        return ""
    if key == "standard_start_time":
        if hasattr(value, "strftime"):
            return value.strftime("%H:%M")
        text = str(value).strip()
        if len(text) >= 5 and text[2] == ":":
            return text[:5]
        return text
    if key in ("overtime_after_hours", "overtime_multiplier"):
        return format(Decimal(str(value)).quantize(Decimal("0.0001")), "f")
    if key in ("rounding_increment_minutes", "break_deduction_minutes", "break_deduction_after_minutes"):
        return str(int(value))
    return str(value).strip()


def _policy_snapshots_equal(current: dict, stored: dict | None) -> bool:
    stored_map = stored if isinstance(stored, dict) else {}
    for key in _POLICY_SNAPSHOT_COMPARE_KEYS:
        if _normalize_policy_snapshot_field(key, current.get(key)) != _normalize_policy_snapshot_field(
            key,
            stored_map.get(key),
        ):
            return False
    return True


def _resolve_current_cis_payment_mode_for_item(item: PayrollItem, profile: object | None) -> str:
    source = _stored_payment_mode_source_or_none(getattr(item, "payment_mode_source", None))
    if item.status == "pending" and source == "manual":
        return normalize_payroll_payment_mode(item.payment_mode)
    profile_mode = _stored_payment_mode_or_none(getattr(profile, "payment_mode", None) if profile else None)
    return profile_mode or "net_payment"


def _payroll_item_inputs_stale(
    db_session: Session,
    *,
    item: PayrollItem,
    profile: object | None,
    company_default_tax: float | None,
    workplace_tax: float | None,
    company_policy: object,
    week_start: date,
) -> bool:
    """Read-only: true when live CIS inputs differ from values snapshotted on the payroll item."""
    current_hourly = None
    if profile is not None and getattr(profile, "hourly_rate", None) is not None:
        current_hourly = Decimal(str(profile.hourly_rate))
    if not _decimal_snapshots_equal(current_hourly, item.hourly_rate_snapshot):
        return True

    current_tax = resolve_effective_tax_rate_percent(profile, company_default_tax, workplace_tax)
    if not _decimal_snapshots_equal(current_tax, item.tax_rate_snapshot):
        return True

    current_mode = _resolve_current_cis_payment_mode_for_item(item, profile)
    if current_mode != normalize_payroll_payment_mode(item.payment_mode):
        return True

    current_policy_snapshot = policy_snapshot_dict(company_policy)
    if not _policy_snapshots_equal(current_policy_snapshot, item.policy_snapshot):
        return True

    live_seconds = sum_rounded_seconds_payroll_week(
        db_session,
        company_id=item.company_id,
        user_id=item.user_id,
        week_start=week_start,
        policy=company_policy,
    )
    if int(live_seconds) != int(item.rounded_total_seconds or 0):
        return True

    return False


def _payroll_period_has_stale_item_inputs(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    policy: object,
    week_start: date,
    all_items: list[PayrollItem],
) -> bool:
    company = get_company_by_id(db_session, company_id)
    default_tax = float(company.default_tax_rate) if company is not None and company.default_tax_rate is not None else None
    workplace_tax = first_workplace_tax(db_session, company_id)
    for item in all_items:
        profile = get_employee_profile_by_user_id(db_session, item.user_id)
        if _payroll_item_inputs_stale(
            db_session,
            item=item,
            profile=profile,
            company_default_tax=default_tax,
            workplace_tax=workplace_tax,
            company_policy=policy,
            week_start=week_start,
        ):
            return True
    return False


def _build_report_alerts(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    policy,
    week_start: date,
    period: PayrollPeriod | None,
    all_items: list[PayrollItem],
) -> PayrollReportAlerts:
    week_start_utc, week_end_utc = week_bounds_utc(policy, week_start)
    open_n = count_open_shifts_started_in_week(
        db_session,
        company_id=company_id,
        week_start_utc=week_start_utc,
        week_end_utc=week_end_utc,
    )
    pending = sum(1 for i in all_items if _item_has_actionable_pending_payroll_value(i))
    rate_missing = _missing_profile_hourly_rate_count(db_session, all_items)
    missing_setup = sum(1 for i in all_items if _item_missing_required_payroll_setup(i))
    utr_missing, nino_missing = _missing_tax_identifier_counts(db_session, all_items)
    zero_hours = sum(1 for i in all_items if i.rounded_total_seconds == 0)
    has_items = len(all_items) > 0
    not_calculated = period is None or (period.calculated_at is None and not has_items)
    needs_recalc = bool(period is not None and period.calculated_at is None and has_items)
    if period is not None and period.calculated_at is not None:
        max_shift_updated = max_employee_shift_updated_at_in_payroll_week(
            db_session,
            company_id=company_id,
            week_start_utc=week_start_utc,
            week_end_utc=week_end_utc,
        )
        if max_shift_updated is not None and max_shift_updated > period.calculated_at:
            needs_recalc = True
        if not needs_recalc and has_items:
            needs_recalc = _payroll_period_has_stale_item_inputs(
                db_session,
                company_id=company_id,
                policy=policy,
                week_start=week_start,
                all_items=all_items,
            )
    approved_n = sum(1 for i in all_items if i.status == "approved")
    paid_n = sum(1 for i in all_items if i.status == "paid")
    can_auto = not_calculated and approved_n == 0 and paid_n == 0
    return PayrollReportAlerts(
        pending_approval_count=pending,
        open_shifts_started_in_week_count=open_n,
        rate_missing_employees_count=rate_missing,
        missing_payroll_setup_employees_count=missing_setup,
        utr_missing_employees_count=utr_missing,
        nino_missing_employees_count=nino_missing,
        zero_rounded_hours_employees_count=zero_hours,
        payroll_period_not_calculated=not_calculated,
        payroll_needs_recalculation=needs_recalc,
        can_auto_recalculate=can_auto,
    )


def _payroll_period_needs_recalculation(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
    period: PayrollPeriod,
    all_items: list[PayrollItem] | None = None,
) -> bool:
    items = all_items if all_items is not None else list_items_for_period(db_session, period.id)
    policy = ensure_company_time_policy(db_session, company_id)
    return _build_report_alerts(
        db_session,
        company_id=company_id,
        policy=policy,
        week_start=week_start,
        period=period,
        all_items=items,
    ).payroll_needs_recalculation


def _assert_payroll_period_not_stale_for_approval(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    period: PayrollPeriod,
    all_items: list[PayrollItem] | None = None,
) -> None:
    if _payroll_period_needs_recalculation(
        db_session,
        company_id=company_id,
        week_start=period.week_start,
        period=period,
        all_items=all_items,
    ):
        raise PayrollItemStateError("Payroll needs recalculation before approval.")


def _assert_payroll_items_ready_for_approval(items: list[PayrollItem]) -> None:
    pending_items = [item for item in items if getattr(item, "status", None) == "pending"]
    missing_rate = sum(1 for item in pending_items if _item_missing_hourly_rate_for_approval(item))
    if missing_rate > 0:
        raise PayrollItemStateError(
            f"Cannot approve payroll: {missing_rate} pending row"
            f"{'' if missing_rate == 1 else 's'} missing or invalid hourly rate."
        )
    missing_setup = sum(1 for item in pending_items if _item_missing_required_payroll_setup(item))
    if missing_setup > 0:
        raise PayrollItemStateError(
            f"Cannot approve payroll: {missing_setup} pending row"
            f"{'' if missing_setup == 1 else 's'} missing payroll/CIS setup."
        )


def _compute_late_unpaid_employees(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
    period: PayrollPeriod,
    all_items: list[PayrollItem],
    policy,
) -> tuple[list[PayrollLateUnpaidEmployee], int, int, int]:
    """Employees with paid items and completed shifts after paid_at not yet reserved on a pending adjustment.

    Returns (blocks, total_rounded_seconds, detected_shift_count, payable_shift_count).
    """
    company = get_company_by_id(db_session, company_id)
    default_tax = float(company.default_tax_rate) if company is not None and company.default_tax_rate is not None else None
    workplace_tax = first_workplace_tax(db_session, company_id)
    week_start_utc, week_end_utc = week_bounds_utc(policy, week_start)
    paid_by_user: dict[uuid.UUID, tuple[datetime, uuid.UUID]] = {}
    for it in all_items:
        if it.status != "paid" or it.paid_at is None:
            continue
        cur = paid_by_user.get(it.user_id)
        if cur is None or it.paid_at > cur[0]:
            paid_by_user[it.user_id] = (it.paid_at, it.id)
    blocks: list[PayrollLateUnpaidEmployee] = []
    total_seconds = 0
    detected_shift_total = 0
    payable_shift_total = 0
    ot_mult = Decimal(str(policy.overtime_multiplier))
    for user_id, (paid_cutoff, ref_item_id) in paid_by_user.items():
        reserved = reserved_late_shift_ids_for_user_period(all_items, user_id)
        rows = list_time_shifts_for_payroll_week(
            db_session,
            company_id=company_id,
            subject_user_id=user_id,
            week_start_utc=week_start_utc,
            week_end_utc=week_end_utc,
        )
        late_rows: list[PayrollLateShiftRow] = []
        sum_sec = 0
        for shift, location, _owner, profile in rows:
            if not shift_completed_after_paid_cutoff(shift, paid_cutoff):
                continue
            if shift.id in reserved:
                continue
            pol = effective_time_policy_for_shift(db_session, shift, location)
            profile_early = bool(profile.early_access_enabled) if profile is not None else False
            early_access = effective_early_access_for_shift(
                db_session, location, profile_early_access=profile_early
            )
            metrics = compute_shift_metrics(
                clock_in_at_utc=shift.clock_in_at,
                clock_out_at_utc=shift.clock_out_at,
                break_seconds_tracked=int(shift.break_seconds or 0),
                early_access_enabled=early_access,
                policy=pol,
            )
            rs = int(metrics.rounded_seconds or 0)
            late_rows.append(
                PayrollLateShiftRow(
                    shift_id=shift.id,
                    clock_in_at=shift.clock_in_at,
                    clock_out_at=shift.clock_out_at,
                    rounded_seconds=rs,
                    reason="completed_after_paid",
                    reference_paid_item_id=ref_item_id,
                )
            )
            sum_sec += rs
        if not late_rows:
            continue
        detected_shift_total += len(late_rows)
        payable_shift_total += sum(1 for r in late_rows if r.rounded_seconds > 0)
        profile = get_employee_profile_by_user_id(db_session, user_id)
        hourly = None
        if profile is not None and profile.hourly_rate is not None:
            hourly = Decimal(str(profile.hourly_rate))
        tax_pct = resolve_effective_tax_rate_percent(profile, default_tax, workplace_tax)
        ref_item = next((it for it in all_items if it.id == ref_item_id), None)
        pay_mode = (
            normalize_payroll_payment_mode(ref_item.payment_mode) if ref_item is not None else "net_payment"
        )
        by_day: dict[date, int] = {}
        for row in late_rows:
            wd = work_date_in_policy(row.clock_in_at, policy)
            by_day[wd] = by_day.get(wd, 0) + row.rounded_seconds
        reg_s, ot_s = split_regular_overtime_daily_by_work_date(by_day, policy.overtime_after_hours)
        bundle = compute_money_bundle(
            regular_seconds=reg_s,
            overtime_seconds=ot_s,
            hourly_rate=hourly,
            overtime_multiplier=ot_mult,
            tax_rate_percent=tax_pct,
            other_deductions=Decimal(0),
            payment_mode=pay_mode,
        )
        eg = Decimal(str(bundle["gross_amount"])) if bundle["gross_amount"] is not None else None
        en = Decimal(str(bundle["net_amount"])) if bundle["net_amount"] is not None else None
        et = Decimal(str(bundle["tax_amount"])) if bundle["tax_amount"] is not None else None
        email, name, _jt = _employee_display(db_session, user_id)
        blocks.append(
            PayrollLateUnpaidEmployee(
                user_id=user_id,
                employee_email=email,
                employee_name=name,
                total_late_rounded_seconds=sum_sec,
                shifts=late_rows,
                estimated_gross_amount=eg,
                estimated_net_amount=en,
                estimated_cis_tax_amount=et,
            )
        )
        total_seconds += sum_sec
    return blocks, total_seconds, detected_shift_total, payable_shift_total


def _late_shift_rounded_entries_after_paid_cutoff(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    week_start: date,
    policy,
    user_id: uuid.UUID,
    paid_cutoff: datetime,
    reserved_ids: set[uuid.UUID],
) -> list[tuple[uuid.UUID, int, date]]:
    """Completed shifts in the payroll week after paid_cutoff, excluding IDs reserved on pending rows."""
    week_start_utc, week_end_utc = week_bounds_utc(policy, week_start)
    rows = list_time_shifts_for_payroll_week(
        db_session,
        company_id=company_id,
        subject_user_id=user_id,
        week_start_utc=week_start_utc,
        week_end_utc=week_end_utc,
    )
    out: list[tuple[uuid.UUID, int, date]] = []
    for shift, location, _owner, profile in rows:
        if not shift_completed_after_paid_cutoff(shift, paid_cutoff):
            continue
        if shift.id in reserved_ids:
            continue
        pol = effective_time_policy_for_shift(db_session, shift, location)
        profile_early = bool(profile.early_access_enabled) if profile is not None else False
        early_access = effective_early_access_for_shift(
            db_session, location, profile_early_access=profile_early
        )
        metrics = compute_shift_metrics(
            clock_in_at_utc=shift.clock_in_at,
            clock_out_at_utc=shift.clock_out_at,
            break_seconds_tracked=int(shift.break_seconds or 0),
            early_access_enabled=early_access,
            policy=pol,
        )
        rs = int(metrics.rounded_seconds or 0)
        out.append((shift.id, rs, work_date_in_policy(shift.clock_in_at, policy)))
    return out


def item_to_response(db_session: Session, item: PayrollItem) -> PayrollItemResponse:
    email, name, job_title = _employee_display(db_session, item.user_id)
    return PayrollItemResponse(
        id=item.id,
        period_id=item.period_id,
        user_id=item.user_id,
        company_id=item.company_id,
        employee_email=email,
        employee_name=name,
        employee_job_title=job_title,
        regular_seconds=item.regular_seconds,
        overtime_seconds=item.overtime_seconds,
        rounded_total_seconds=item.rounded_total_seconds,
        hourly_rate_snapshot=_decimal_or_none(item.hourly_rate_snapshot),
        tax_rate_snapshot=_decimal_or_none(item.tax_rate_snapshot),
        overtime_multiplier_snapshot=_decimal_or_none(item.overtime_multiplier_snapshot),
        gross_amount=_decimal_or_none(item.gross_amount),
        tax_amount=_decimal_or_none(item.tax_amount),
        net_amount=_decimal_or_none(item.net_amount),
        other_deductions_amount=Decimal(str(item.other_deductions_amount or 0)),
        display_tax_amount=_decimal_or_none(item.display_tax_amount),
        display_net_amount=_decimal_or_none(item.display_net_amount),
        payment_mode=_stored_payment_mode_or_none(item.payment_mode),
        payment_mode_label=_payment_mode_label_for_item(item),
        notes=item.notes,
        policy_snapshot=item.policy_snapshot or {},
        status=item.status,
        approved_at=item.approved_at,
        approved_by_user_id=item.approved_by_user_id,
        paid_at=item.paid_at,
        paid_by_user_id=item.paid_by_user_id,
        rate_missing=item.rate_missing,
    )


def _summarize_period(
    db_session: Session,
    period: PayrollPeriod,
    items: list[PayrollItem],
) -> PayrollPeriodSummary:
    pending = sum(1 for i in items if i.status == "pending")
    approved = sum(1 for i in items if i.status == "approved")
    paid = sum(1 for i in items if i.status == "paid")
    tr = tor = tt = tg = tn = Decimal(0)
    for i in items:
        tr += i.regular_seconds
        tor += i.overtime_seconds
        tt += i.rounded_total_seconds
        if i.gross_amount is not None:
            tg += Decimal(str(i.gross_amount))
        if i.display_net_amount is not None:
            tn += Decimal(str(i.display_net_amount))
        elif i.net_amount is not None:
            tn += Decimal(str(i.net_amount))
    total_tax = Decimal(0)
    for i in items:
        eff = _effective_tax_amount_for_item(i)
        if eff is not None:
            total_tax += eff
    other_sum = sum(Decimal(str(i.other_deductions_amount or 0)) for i in items)
    return PayrollPeriodSummary(
        id=period.id,
        company_id=period.company_id,
        week_start=period.week_start,
        timezone_name=period.timezone_name,
        calculated_at=period.calculated_at,
        calculated_by_user_id=period.calculated_by_user_id,
        total_items=len(items),
        pending_count=pending,
        approved_count=approved,
        paid_count=paid,
        total_regular_seconds=tr,
        total_overtime_seconds=tor,
        total_rounded_seconds=tt,
        total_gross=tg if items else None,
        total_tax=total_tax if items else None,
        total_net=tn if items else None,
        total_other_deductions=other_sum,
    )


def get_payroll_report(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date,
    user_id: uuid.UUID | None = None,
    auto_recalculate_if_safe: bool = True,
    _auto_recalc_depth: int = 0,
) -> PayrollReportResponse:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    policy = ensure_company_time_policy(db_session, company_id)
    period = get_period_by_company_week(db_session, company_id, week_start)
    if period is None:
        alerts = _build_report_alerts(
            db_session,
            company_id=company_id,
            policy=policy,
            week_start=week_start,
            period=None,
            all_items=[],
        )
        if (
            auto_recalculate_if_safe
            and _auto_recalc_depth == 0
            and alerts.can_auto_recalculate
        ):
            recalculate_payroll(db_session, actor, company_id=company_id, week_start=week_start)
            inner = get_payroll_report(
                db_session,
                actor,
                company_id=company_id,
                week_start=week_start,
                user_id=user_id,
                auto_recalculate_if_safe=auto_recalculate_if_safe,
                _auto_recalc_depth=_auto_recalc_depth + 1,
            )
            return inner.model_copy(update={"payroll_auto_recalculated": True})
        empty = PayrollPeriodSummary(
            id=uuid.UUID("00000000-0000-0000-0000-000000000000"),
            company_id=company_id,
            week_start=week_start,
            timezone_name=policy.timezone_name,
            calculated_at=None,
            calculated_by_user_id=None,
            total_items=0,
            pending_count=0,
            approved_count=0,
            paid_count=0,
            total_regular_seconds=0,
            total_overtime_seconds=0,
            total_rounded_seconds=0,
            total_gross=None,
            total_tax=None,
            total_net=None,
            total_other_deductions=Decimal(0),
        )
        split = _build_pay_split([])
        acct_overlap = _accounting_export_overlaps_payroll_week(
            db_session, company_id=company_id, week_start=week_start
        )
        return PayrollReportResponse(
            period=empty,
            items=[],
            alerts=alerts,
            split=split,
            has_late_unpaid_shifts=False,
            late_shift_count=0,
            late_shift_count_detected=0,
            late_shift_count_payable=0,
            late_unpaid_total_rounded_seconds=0,
            has_payable_late_unpaid_shifts=False,
            late_unpaid_employees=[],
            accounting_payroll_export_overlaps=acct_overlap,
        )

    all_items = list_items_for_period(db_session, period.id)
    if user_id is not None:
        target = get_user_by_id(db_session, user_id)
        if (
            target is None
            or target.company_id != company_id
            or target.system_role != SystemRole.EMPLOYEE
        ):
            raise PayrollError("Invalid employee filter.")
        display_items = [i for i in all_items if i.user_id == user_id]
    else:
        display_items = all_items

    alerts = _build_report_alerts(
        db_session,
        company_id=company_id,
        policy=policy,
        week_start=week_start,
        period=period,
        all_items=all_items,
    )
    if (
        auto_recalculate_if_safe
        and _auto_recalc_depth == 0
        and alerts.can_auto_recalculate
    ):
        recalculate_payroll(db_session, actor, company_id=company_id, week_start=week_start)
        inner = get_payroll_report(
            db_session,
            actor,
            company_id=company_id,
            week_start=week_start,
            user_id=user_id,
            auto_recalculate_if_safe=auto_recalculate_if_safe,
            _auto_recalc_depth=_auto_recalc_depth + 1,
        )
        return inner.model_copy(update={"payroll_auto_recalculated": True})
    split = _build_pay_split(all_items)
    late_employees, late_secs, late_n, late_payable_n = _compute_late_unpaid_employees(
        db_session,
        company_id=company_id,
        week_start=week_start,
        period=period,
        all_items=all_items,
        policy=policy,
    )
    acct_overlap = _accounting_export_overlaps_payroll_week(
        db_session, company_id=company_id, week_start=period.week_start
    )
    leave_rows = _payroll_approved_leave_rows(
        db_session,
        company_id=company_id,
        week_start=week_start,
        filter_user_id=user_id,
    )
    has_payable_late = late_payable_n > 0 or late_secs > 0
    return PayrollReportResponse(
        period=_summarize_period(db_session, period, all_items),
        items=[item_to_response(db_session, i) for i in display_items],
        alerts=alerts,
        split=split,
        has_late_unpaid_shifts=late_n > 0,
        late_shift_count=late_n,
        late_shift_count_detected=late_n,
        late_shift_count_payable=late_payable_n,
        late_unpaid_total_rounded_seconds=late_secs,
        has_payable_late_unpaid_shifts=has_payable_late,
        late_unpaid_employees=late_employees,
        accounting_payroll_export_overlaps=acct_overlap,
        approved_leave_in_week=leave_rows,
    )


def get_payroll_month_summary(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    year: int,
    month: int,
) -> PayrollMonthSummaryResponse:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    if month < 1 or month > 12 or year < 2000 or year > 2100:
        raise PayrollError("Invalid month or year.")
    periods = list_periods_for_company_month(db_session, company_id, year=year, month=month)
    all_items: list[PayrollItem] = []
    user_ids: set[uuid.UUID] = set()
    for p in periods:
        for row in list_items_for_period(db_session, p.id):
            all_items.append(row)
            user_ids.add(row.user_id)
    tr = tor = tt = 0
    tg = tn = total_tax = Decimal(0)
    other_sum = Decimal(0)
    has_gross = has_net = has_tax = False
    for i in all_items:
        tr += i.regular_seconds
        tor += i.overtime_seconds
        tt += i.rounded_total_seconds
        if i.gross_amount is not None:
            tg += Decimal(str(i.gross_amount))
            has_gross = True
        if i.display_net_amount is not None:
            tn += Decimal(str(i.display_net_amount))
            has_net = True
        elif i.net_amount is not None:
            tn += Decimal(str(i.net_amount))
            has_net = True
        eff_tax = _effective_tax_amount_for_item(i)
        if eff_tax is not None:
            total_tax += eff_tax
            has_tax = True
        other_sum += Decimal(str(i.other_deductions_amount or 0))
    return PayrollMonthSummaryResponse(
        company_id=company_id,
        year=year,
        month=month,
        payroll_weeks=len(periods),
        distinct_employees=len(user_ids),
        total_regular_seconds=tr,
        total_overtime_seconds=tor,
        total_rounded_seconds=tt,
        total_gross=tg if has_gross else None,
        total_tax=total_tax if has_tax else None,
        total_net=tn if has_net else None,
        total_other_deductions=other_sum,
        total_days=None,
    )


def recalculate_payroll(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date,
) -> PayrollReportResponse:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    company = get_company_by_id(db_session, company_id)
    if company is None:
        raise PayrollError("Company not found.")
    policy = ensure_company_time_policy(db_session, company_id)
    workplace_tax = first_workplace_tax(db_session, company_id)
    default_tax = float(company.default_tax_rate) if company.default_tax_rate is not None else None

    period = get_period_by_company_week(db_session, company_id, week_start)
    if period is not None and period_has_paid_item(db_session, period.id):
        raise PayrollPaidBlockingError(
            "Cannot recalculate: this period contains paid payroll items.",
        )
    if period is not None and period_has_approved_item(db_session, period.id):
        raise PayrollApprovedBlockingError(
            "Payroll is approved. Unlock it before recalculating.",
        )
    if period is None:
        period = PayrollPeriod(
            company_id=company_id,
            week_start=week_start,
            timezone_name=policy.timezone_name,
        )
    else:
        period.timezone_name = policy.timezone_name
    period = save_period(db_session, period)
    pending_manual_modes: dict[uuid.UUID, str] = {}
    for it in list_items_for_period(db_session, period.id):
        pending_mode = _stored_payment_mode_or_none(it.payment_mode)
        payment_mode_source = _stored_payment_mode_source_or_none(getattr(it, "payment_mode_source", None))
        if it.status == "pending" and pending_mode is not None and payment_mode_source == "manual":
            pending_manual_modes[it.user_id] = pending_mode
    delete_pending_items_for_period(db_session, period.id)

    employees = list_cis_employee_users_for_company(db_session, company_id)
    now = datetime.now(timezone.utc)
    ot_mult = Decimal(str(policy.overtime_multiplier))

    for emp in employees:
        profile = get_employee_profile_by_user_id(db_session, emp.id)
        reg_s, ot_s, total_r = regular_overtime_seconds_payroll_week(
            db_session,
            company_id=company_id,
            user_id=emp.id,
            week_start=week_start,
            policy=policy,
        )
        hourly = None
        if profile is not None and profile.hourly_rate is not None:
            hourly = Decimal(str(profile.hourly_rate))
        tax_pct = resolve_effective_tax_rate_percent(profile, default_tax, workplace_tax)
        other_d = Decimal(0)
        profile_mode = _stored_payment_mode_or_none(getattr(profile, "payment_mode", None))
        manual_mode = pending_manual_modes.get(emp.id)
        pay_mode = manual_mode or profile_mode or "net_payment"
        payment_mode_source = "manual" if manual_mode is not None else "profile"
        bundle = compute_money_bundle(
            regular_seconds=reg_s,
            overtime_seconds=ot_s,
            hourly_rate=hourly,
            overtime_multiplier=ot_mult,
            tax_rate_percent=tax_pct,
            other_deductions=other_d,
            payment_mode=pay_mode,
        )
        snap = policy_snapshot_dict(policy)
        item = PayrollItem(
            period_id=period.id,
            user_id=emp.id,
            company_id=company_id,
            regular_seconds=reg_s,
            overtime_seconds=ot_s,
            rounded_total_seconds=total_r,
            hourly_rate_snapshot=float(hourly) if hourly is not None else None,
            tax_rate_snapshot=float(tax_pct) if tax_pct is not None else None,
            overtime_multiplier_snapshot=float(ot_mult),
            gross_amount=float(bundle["gross_amount"]) if bundle["gross_amount"] is not None else None,
            tax_amount=float(bundle["tax_amount"]) if bundle["tax_amount"] is not None else None,
            net_amount=float(bundle["net_amount"]) if bundle["net_amount"] is not None else None,
            other_deductions_amount=float(other_d),
            display_tax_amount=float(bundle["display_tax_amount"]) if bundle["display_tax_amount"] is not None else None,
            display_net_amount=float(bundle["display_net_amount"]) if bundle["display_net_amount"] is not None else None,
            policy_snapshot=snap,
            status="pending",
            rate_missing=bool(bundle["rate_missing"]),
            payment_mode=pay_mode,
            payment_mode_source=payment_mode_source,
        )
        save_item(db_session, item)

    period.calculated_at = now
    period.calculated_by_user_id = actor.id
    save_period(db_session, period)

    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_recalculated",
        entity_type="payroll_period",
        entity_id=str(period.id),
        company_id=company_id,
        details={"week_start": str(week_start)},
    )

    items = list_items_for_period(db_session, period.id)
    policy = ensure_company_time_policy(db_session, company_id)
    alerts = _build_report_alerts(
        db_session,
        company_id=company_id,
        policy=policy,
        week_start=week_start,
        period=period,
        all_items=items,
    )
    split = _build_pay_split(items)
    late_employees, late_secs, late_n, late_payable_n = _compute_late_unpaid_employees(
        db_session,
        company_id=company_id,
        week_start=week_start,
        period=period,
        all_items=items,
        policy=policy,
    )
    acct_overlap = _accounting_export_overlaps_payroll_week(
        db_session, company_id=company_id, week_start=period.week_start
    )
    leave_rows = _payroll_approved_leave_rows(
        db_session,
        company_id=company_id,
        week_start=week_start,
        filter_user_id=None,
    )
    has_payable_late = late_payable_n > 0 or late_secs > 0
    return PayrollReportResponse(
        period=_summarize_period(db_session, period, items),
        items=[item_to_response(db_session, i) for i in items],
        alerts=alerts,
        split=split,
        has_late_unpaid_shifts=late_n > 0,
        late_shift_count=late_n,
        late_shift_count_detected=late_n,
        late_shift_count_payable=late_payable_n,
        late_unpaid_total_rounded_seconds=late_secs,
        has_payable_late_unpaid_shifts=has_payable_late,
        late_unpaid_employees=late_employees,
        accounting_payroll_export_overlaps=acct_overlap,
        approved_leave_in_week=leave_rows,
    )


def patch_payroll_item(
    db_session: Session,
    actor: User,
    item_id: uuid.UUID,
    request: PayrollItemPatchRequest,
) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, item.company_id)

    if item.status == "paid":
        money_touched = (
            request.payment_mode is not None
            or request.other_deductions_amount is not None
            or request.display_tax_amount is not None
            or request.display_net_amount is not None
        )
        if money_touched:
            raise PayrollError("Paid payroll rows are locked; only notes may be edited.")
        if request.notes is not None:
            item.notes = request.notes
        item.updated_at = datetime.now(timezone.utc)
        update_item(db_session, item)
        create_internal_audit_event(
            db_session=db_session,
            actor=actor,
            action="payroll_item_edited",
            entity_type="payroll_item",
            entity_id=str(item.id),
            company_id=item.company_id,
            details={"paid_locked": True},
        )
        return item_to_response(db_session, item)

    existing_payment_mode = _stored_payment_mode_or_none(item.payment_mode)

    if request.notes is not None:
        item.notes = request.notes
    if request.payment_mode is not None:
        requested_payment_mode = normalize_payroll_payment_mode(request.payment_mode)
        item.payment_mode = requested_payment_mode
        if requested_payment_mode != existing_payment_mode:
            item.payment_mode_source = "manual"
    else:
        item.payment_mode = normalize_payroll_payment_mode(item.payment_mode)
    if request.other_deductions_amount is not None:
        item.other_deductions_amount = float(request.other_deductions_amount)
    if request.display_tax_amount is not None:
        item.display_tax_amount = float(request.display_tax_amount)
    if request.display_net_amount is not None:
        item.display_net_amount = float(request.display_net_amount)

    _apply_payroll_item_money_after_patch(item, request)

    item.updated_at = datetime.now(timezone.utc)
    update_item(db_session, item)

    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_item_edited",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={},
    )
    return item_to_response(db_session, item)


def approve_item(db_session: Session, actor: User, item_id: uuid.UUID) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, item.company_id)
    if item.status != "pending":
        raise PayrollItemStateError("Only pending rows can be approved.")
    period = db_session.get(PayrollPeriod, item.period_id)
    if period is None:
        raise PayrollError("Payroll period not found.")
    _assert_payroll_period_not_stale_for_approval(
        db_session,
        company_id=item.company_id,
        period=period,
    )
    _assert_payroll_items_ready_for_approval([item])
    item.status = "approved"
    item.approved_at = datetime.now(timezone.utc)
    item.approved_by_user_id = actor.id
    update_item(db_session, item)
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_item_approved",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={},
    )
    return item_to_response(db_session, item)


def unlock_item(db_session: Session, actor: User, item_id: uuid.UUID) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, item.company_id)
    if item.status == "paid":
        raise PayrollItemStateError("Paid rows cannot be unlocked.")
    if item.status != "approved":
        raise PayrollItemStateError("Only approved rows can be unlocked.")
    item.status = "pending"
    item.approved_at = None
    item.approved_by_user_id = None
    update_item(db_session, item)
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_item_unlocked",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={},
    )
    return item_to_response(db_session, item)


def mark_paid_item(db_session: Session, actor: User, item_id: uuid.UUID) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, item.company_id)
    if item.status != "approved":
        raise PayrollItemStateError("Only approved rows can be marked paid.")
    item.status = "paid"
    item.paid_at = datetime.now(timezone.utc)
    item.paid_by_user_id = actor.id
    update_item(db_session, item)
    record_payroll_paid(
        db_session,
        company_id=item.company_id,
        payroll_item_id=item.id,
        employee_user_id=item.user_id,
    )
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_item_marked_paid",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={},
    )
    return item_to_response(db_session, item)


def undo_paid_item(
    db_session: Session,
    actor: User,
    item_id: uuid.UUID,
    request: PayrollUndoPaidRequest,
) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    if not request.confirm:
        raise PayrollError("confirm must be true to undo paid status.")
    item = get_item_by_id(db_session, item_id)
    if item is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, item.company_id)
    if item.status != "paid":
        raise PayrollItemStateError("Only paid rows can be moved back to approved.")
    period = db_session.get(PayrollPeriod, item.period_id)
    if period is None:
        raise PayrollError("Payroll period not found.")
    overlap = _accounting_export_overlaps_payroll_week(
        db_session,
        company_id=item.company_id,
        week_start=period.week_start,
    )
    if overlap and not request.acknowledge_accounting_export:
        raise PayrollError(
            "An accounting payroll export overlaps this week. "
            "Set acknowledge_accounting_export to true if you still want to undo paid."
        )
    prev_paid_at = item.paid_at.isoformat() if item.paid_at else None
    prev_paid_by = str(item.paid_by_user_id) if item.paid_by_user_id else None
    item.status = "approved"
    item.paid_at = None
    item.paid_by_user_id = None
    item.updated_at = datetime.now(timezone.utc)
    update_item(db_session, item)
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.payment_undone",
        entity_type="payroll_item",
        entity_id=str(item.id),
        company_id=item.company_id,
        details={
            "item_id": str(item.id),
            "period_id": str(item.period_id),
            "user_id": str(item.user_id),
            "company_id": str(item.company_id),
            "week_start": str(period.week_start),
            "reason": request.reason.strip(),
            "actor_user_id": str(actor.id),
            "previous_paid_at": prev_paid_at,
            "previous_paid_by_user_id": prev_paid_by,
            "accounting_export_overlap": overlap,
        },
    )
    return item_to_response(db_session, item)


def create_late_shift_adjustment_from_paid_item(
    db_session: Session,
    actor: User,
    paid_item_id: uuid.UUID,
    request: PayrollLateAdjustmentRequest,
) -> PayrollItemResponse:
    assert_payroll_admin_or_administrator(actor)
    if not request.confirm:
        raise PayrollError("confirm must be true to create an adjustment.")
    paid_ref = get_item_by_id(db_session, paid_item_id)
    if paid_ref is None:
        raise PayrollError("Payroll item not found.")
    assert_payroll_company_scope(actor, paid_ref.company_id)
    if paid_ref.status != "paid" or paid_ref.paid_at is None:
        raise PayrollItemStateError(
            "Create adjustment from a paid payroll row that has a paid timestamp.",
        )
    period = db_session.get(PayrollPeriod, paid_ref.period_id)
    if period is None:
        raise PayrollError("Payroll period not found.")
    policy = ensure_company_time_policy(db_session, paid_ref.company_id)
    all_items = list_items_for_period(db_session, period.id)
    reserved = reserved_late_shift_ids_for_user_period(all_items, paid_ref.user_id)
    candidates = _late_shift_rounded_entries_after_paid_cutoff(
        db_session,
        company_id=paid_ref.company_id,
        week_start=period.week_start,
        policy=policy,
        user_id=paid_ref.user_id,
        paid_cutoff=paid_ref.paid_at,
        reserved_ids=reserved,
    )
    cand_map: dict[uuid.UUID, tuple[int, date]] = {
        sid: (sec, wd) for sid, sec, wd in candidates
    }
    if request.shift_ids:
        wanted = set(request.shift_ids)
        if not wanted.issubset(cand_map.keys()):
            raise PayrollError("One or more shift_ids are not uncovered late shifts for this paid row.")
        selected = [(sid, cand_map[sid][0], cand_map[sid][1]) for sid in wanted]
    else:
        selected = list(candidates)
    if not selected:
        raise PayrollError("No late unpaid shifts remain for an adjustment.")
    if all(sec <= 0 for _sid, sec, _wd in selected):
        if len(selected) == 1:
            raise PayrollError(
                "No payable late hours were found. The detected shift has zero payroll-rounded time.",
            )
        raise PayrollError(
            "No payable late hours were found. The detected shifts have zero payroll-rounded time.",
        )
    total_r = sum(sec for _sid, sec, _wd in selected)
    shift_ids_ordered = [sid for sid, _sec, _wd in selected]
    by_day: dict[date, int] = {}
    for _sid, sec, work_date in selected:
        by_day[work_date] = by_day.get(work_date, 0) + sec
    reg_s, ot_s = split_regular_overtime_daily_by_work_date(by_day, policy.overtime_after_hours)
    pay_mode = normalize_payroll_payment_mode(paid_ref.payment_mode)
    hourly = _decimal_or_none(paid_ref.hourly_rate_snapshot)
    tax_pct = _decimal_or_none(paid_ref.tax_rate_snapshot)
    ot_mult = _decimal_or_none(paid_ref.overtime_multiplier_snapshot) or Decimal(1)
    other_d = Decimal(0)
    bundle = compute_money_bundle(
        regular_seconds=reg_s,
        overtime_seconds=ot_s,
        hourly_rate=hourly,
        overtime_multiplier=ot_mult,
        tax_rate_percent=tax_pct,
        other_deductions=other_d,
        payment_mode=pay_mode,
    )
    snap = paid_ref.policy_snapshot if paid_ref.policy_snapshot else policy_snapshot_dict(policy)
    base_note = "Adjustment for late completed shifts after payroll was paid"
    notes_val = append_late_shift_ids_marker(base_note, shift_ids_ordered)
    new_item = PayrollItem(
        period_id=period.id,
        user_id=paid_ref.user_id,
        company_id=paid_ref.company_id,
        regular_seconds=reg_s,
        overtime_seconds=ot_s,
        rounded_total_seconds=total_r,
        hourly_rate_snapshot=float(hourly) if hourly is not None else None,
        tax_rate_snapshot=float(tax_pct) if tax_pct is not None else None,
        overtime_multiplier_snapshot=float(ot_mult),
        gross_amount=float(bundle["gross_amount"]) if bundle["gross_amount"] is not None else None,
        tax_amount=float(bundle["tax_amount"]) if bundle["tax_amount"] is not None else None,
        net_amount=float(bundle["net_amount"]) if bundle["net_amount"] is not None else None,
        other_deductions_amount=float(other_d),
        display_tax_amount=float(bundle["display_tax_amount"]) if bundle["display_tax_amount"] is not None else None,
        display_net_amount=float(bundle["display_net_amount"]) if bundle["display_net_amount"] is not None else None,
        policy_snapshot=snap,
        status="pending",
        rate_missing=bool(bundle["rate_missing"]),
        payment_mode=pay_mode,
        payment_mode_source="manual",
        notes=notes_val,
    )
    save_item(db_session, new_item)
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.adjustment_created",
        entity_type="payroll_item",
        entity_id=str(new_item.id),
        company_id=paid_ref.company_id,
        details={
            "new_item_id": str(new_item.id),
            "reference_paid_item_id": str(paid_ref.id),
            "period_id": str(period.id),
            "user_id": str(paid_ref.user_id),
            "company_id": str(paid_ref.company_id),
            "week_start": str(period.week_start),
            "shift_count": len(shift_ids_ordered),
            "rounded_seconds": total_r,
            "actor_user_id": str(actor.id),
        },
    )
    return item_to_response(db_session, new_item)


def approve_all_pending(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date,
) -> PayrollReportResponse:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    period = get_period_by_company_week(db_session, company_id, week_start)
    if period is None:
        raise PayrollError("Payroll period not found. Run recalculate first.")
    items = list_items_for_period(db_session, period.id)
    _assert_payroll_period_not_stale_for_approval(
        db_session,
        company_id=company_id,
        period=period,
        all_items=items,
    )
    _assert_payroll_items_ready_for_approval(items)
    for it in items:
        if it.status == "pending":
            it.status = "approved"
            it.approved_at = datetime.now(timezone.utc)
            it.approved_by_user_id = actor.id
            update_item(db_session, it)
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll_approve_all",
        entity_type="payroll_period",
        entity_id=str(period.id),
        company_id=company_id,
        details={"week_start": str(week_start)},
    )
    items = list_items_for_period(db_session, period.id)
    policy = ensure_company_time_policy(db_session, company_id)
    alerts = _build_report_alerts(
        db_session,
        company_id=company_id,
        policy=policy,
        week_start=week_start,
        period=period,
        all_items=items,
    )
    split = _build_pay_split(items)
    late_employees, late_secs, late_n, late_payable_n = _compute_late_unpaid_employees(
        db_session,
        company_id=company_id,
        week_start=week_start,
        period=period,
        all_items=items,
        policy=policy,
    )
    acct_overlap = _accounting_export_overlaps_payroll_week(
        db_session, company_id=company_id, week_start=period.week_start
    )
    leave_rows = _payroll_approved_leave_rows(
        db_session,
        company_id=company_id,
        week_start=week_start,
        filter_user_id=None,
    )
    has_payable_late = late_payable_n > 0 or late_secs > 0
    return PayrollReportResponse(
        period=_summarize_period(db_session, period, items),
        items=[item_to_response(db_session, i) for i in items],
        alerts=alerts,
        split=split,
        has_late_unpaid_shifts=late_n > 0,
        late_shift_count=late_n,
        late_shift_count_detected=late_n,
        late_shift_count_payable=late_payable_n,
        late_unpaid_total_rounded_seconds=late_secs,
        has_payable_late_unpaid_shifts=has_payable_late,
        late_unpaid_employees=late_employees,
        accounting_payroll_export_overlaps=acct_overlap,
        approved_leave_in_week=leave_rows,
    )


def list_my_pay_history(db_session: Session, actor: User) -> list[PayHistoryEntry]:
    if actor.system_role != SystemRole.EMPLOYEE:
        return []
    items = list_items_for_user_pay_history(db_session, actor.id)
    result: list[PayHistoryEntry] = []
    company_names: dict[uuid.UUID, str] = {}
    for i in items:
        period = db_session.get(PayrollPeriod, i.period_id)
        if period is None:
            continue
        if i.company_id not in company_names:
            co = get_company_by_id(db_session, i.company_id)
            company_names[i.company_id] = co.name if co is not None else "Company"
        eff_cis = _effective_tax_amount_for_item(i)
        eff_net = _effective_net_amount_for_item(i)
        result.append(
            PayHistoryEntry(
                id=i.id,
                company_id=i.company_id,
                week_start=period.week_start,
                week_end=_week_end_display(period.week_start),
                period_id=i.period_id,
                regular_seconds=i.regular_seconds,
                overtime_seconds=i.overtime_seconds,
                rounded_total_seconds=i.rounded_total_seconds,
                gross_amount=_decimal_or_none(i.gross_amount),
                tax_amount=_decimal_or_none(i.tax_amount),
                net_amount=_decimal_or_none(i.net_amount),
                display_tax_amount=_decimal_or_none(i.display_tax_amount),
                display_net_amount=_decimal_or_none(i.display_net_amount),
                other_deductions_amount=Decimal(str(i.other_deductions_amount or 0)),
                status=i.status,
                approved_at=i.approved_at,
                paid_at=i.paid_at,
                rate_missing=i.rate_missing,
                company_name=company_names[i.company_id],
                payment_mode=normalize_payroll_payment_mode(i.payment_mode),
                can_open_payslip=True,
                effective_cis_tax_amount=eff_cis,
                effective_net_amount=eff_net,
                timezone_name=period.timezone_name,
            )
        )
    return result


def _parse_uk_tax_year(tax_year: str) -> tuple[int, int, date, date]:
    cleaned = (tax_year or "").strip().replace("/", "-")
    parts = cleaned.split("-")
    if len(parts) != 2:
        raise PayrollError("tax_year must use YYYY-YYYY format.")
    try:
        start_year = int(parts[0])
        end_year = int(parts[1])
    except ValueError:
        raise PayrollError("tax_year must use YYYY-YYYY format.") from None
    if end_year != start_year + 1:
        raise PayrollError("tax_year must be consecutive years, for example 2025-2026.")
    return start_year, end_year, date(start_year, 4, 6), date(end_year, 4, 5)


def _pay_summary_money(value: Decimal | None) -> Decimal:
    return (value or Decimal(0)).quantize(Decimal("0.01"))


def _pay_summary_tax_status(item: PayrollItem) -> str:
    mode = _stored_payment_mode_or_none(item.payment_mode)
    if mode == "gross_payment":
        return "Gross payment"
    tax_pct = _decimal_or_none(item.tax_rate_snapshot)
    if tax_pct is not None and tax_pct > 0:
        return f"CIS {tax_pct.normalize()}%"
    return _payment_mode_label(mode) if mode is not None else "Not provided"


def _pay_summary_period_label(period: PayrollPeriod) -> str:
    return f"{period.week_start.isoformat()} to {_week_end_display(period.week_start).isoformat()}"


def export_my_tax_year_pay_summary_xlsx(
    db_session: Session,
    actor: User,
    *,
    tax_year: str,
) -> bytes:
    if actor.system_role != SystemRole.EMPLOYEE:
        raise PayrollPermissionError("Only employees can download their own pay summary.")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    start_year, end_year, date_from, date_to = _parse_uk_tax_year(tax_year)
    paid_at_from = datetime.combine(date_from, time.min, tzinfo=timezone.utc)
    paid_at_before = datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=timezone.utc)
    rows = list_paid_items_for_user_tax_year_summary(
        db_session,
        user_id=actor.id,
        paid_at_from=paid_at_from,
        paid_at_before=paid_at_before,
    )

    employee_name = _employee_primary_name(db_session, actor.id)
    ni, utr = _employee_tax_identifiers_for_payroll(db_session, actor.id)
    year_label = f"{start_year}/{end_year}"
    company_names: dict[uuid.UUID, str] = {}

    wb = Workbook()
    payslips = wb.active
    payslips.title = "Payslips"
    companies = wb.create_sheet("Companies")

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    summary_fill = PatternFill("solid", fgColor="F3F4F6")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    money_format = '£#,##0.00;[Red]-£#,##0.00'
    date_format = "yyyy-mm-dd"

    summary_values = {
        "gross": Decimal(0),
        "tax": Decimal(0),
        "admin": Decimal(0),
        "deductions": Decimal(0),
        "additions": Decimal(0),
        "vat": Decimal(0),
        "net": Decimal(0),
    }

    payslips["C2"] = "Name"
    payslips["D2"] = employee_name
    payslips["C3"] = "NI Number"
    payslips["D3"] = ni or ""
    payslips["C4"] = "UTR Number"
    payslips["D4"] = utr or ""
    payslips["C5"] = "Period"
    payslips["D5"] = year_label
    for cell_ref in ("C2", "C3", "C4", "C5", "F2", "F3", "F4", "F5", "F6", "F7"):
        payslips[cell_ref].font = bold
        payslips[cell_ref].fill = summary_fill

    summary_rows = [
        ("F2", "G2", "Total Gross", "gross"),
        ("F3", "G3", "Total Tax / CIS", "tax"),
        ("F4", "G4", "Total Admin Fees", "admin"),
        ("F5", "G5", "Total Deductions", "deductions"),
        ("F6", "G6", "Total Additions", "additions"),
        ("F7", "G7", "Total Take Home", "net"),
    ]
    for label_cell, value_cell, label, _key in summary_rows:
        payslips[label_cell] = label
        payslips[value_cell] = Decimal(0)
        payslips[value_cell].number_format = money_format

    table_header_row = 12
    headers = [
        "№",
        "Period",
        "Payment Date",
        "Company",
        "Tax Status",
        "Total Pay",
        "Tax / CIS",
        "Admin Fee",
        "Deductions",
        "Additions",
        "VAT if applicable",
        "Take Home",
    ]
    for col, label in enumerate(headers, start=2):
        cell = payslips.cell(row=table_header_row, column=col, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for idx, (item, period) in enumerate(rows, start=1):
        company_name = company_names.get(item.company_id)
        if company_name is None:
            company = get_company_by_id(db_session, item.company_id)
            company_name = company.name if company is not None else "Company"
            company_names[item.company_id] = company_name

        gross = _pay_summary_money(_decimal_or_none(item.gross_amount))
        tax = _pay_summary_money(_effective_tax_amount_for_item(item))
        admin_fee = Decimal("0.00")
        other_deductions = _pay_summary_money(_decimal_or_none(item.other_deductions_amount))
        deductions = tax + admin_fee + other_deductions
        additions = Decimal("0.00")
        vat = Decimal("0.00")
        net = _pay_summary_money(_effective_net_amount_for_item(item))

        summary_values["gross"] += gross
        summary_values["tax"] += tax
        summary_values["admin"] += admin_fee
        summary_values["deductions"] += deductions
        summary_values["additions"] += additions
        summary_values["vat"] += vat
        summary_values["net"] += net

        row_idx = table_header_row + idx
        values = [
            idx,
            _pay_summary_period_label(period),
            item.paid_at.date() if item.paid_at is not None else None,
            company_name,
            _pay_summary_tax_status(item),
            gross,
            tax,
            admin_fee,
            deductions,
            additions,
            vat,
            net,
        ]
        for col, value in enumerate(values, start=2):
            cell = payslips.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col == 4:
                cell.number_format = date_format
            if col >= 7:
                cell.number_format = money_format
            cell.alignment = Alignment(vertical="top")

    for label_cell, value_cell, _label, key in summary_rows:
        payslips[value_cell] = summary_values[key]
        payslips[value_cell].number_format = money_format

    totals_row = table_header_row + len(rows) + 2
    payslips.cell(row=totals_row, column=2, value="Totals").font = bold
    for col, key in (
        (7, "gross"),
        (8, "tax"),
        (9, "admin"),
        (10, "deductions"),
        (11, "additions"),
        (12, "vat"),
        (13, "net"),
    ):
        cell = payslips.cell(row=totals_row, column=col, value=summary_values[key])
        cell.font = bold
        cell.border = border
        cell.number_format = money_format

    payslips.freeze_panes = f"B{table_header_row + 1}"
    payslips.auto_filter.ref = f"B{table_header_row}:M{max(table_header_row + 1, table_header_row + len(rows))}"
    for col_idx, width in enumerate([4, 8, 24, 14, 28, 16, 13, 13, 12, 13, 12, 16, 13], start=1):
        payslips.column_dimensions[get_column_letter(col_idx)].width = width

    company_headers = ["№", "Name", "Trading Name", "Reference Number", "Address"]
    for col, label in enumerate(company_headers, start=2):
        cell = companies.cell(row=2, column=col, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.border = border
    for idx, (_company_id, company_name) in enumerate(sorted(company_names.items(), key=lambda row: row[1]), start=1):
        row_idx = 2 + idx
        values = [idx, company_name, "", "", ""]
        for col, value in enumerate(values, start=2):
            cell = companies.cell(row=row_idx, column=col, value=value)
            cell.border = border
    for col_idx, width in enumerate([4, 8, 34, 24, 22, 44], start=1):
        companies.column_dimensions[get_column_letter(col_idx)].width = width

    output = io.BytesIO()
    wb.save(output)

    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.pay_summary_exported",
        entity_type="payroll_item",
        entity_id=None,
        company_id=actor.company_id,
        details={
            "tax_year": f"{start_year}-{end_year}",
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "row_count": len(rows),
        },
    )
    return output.getvalue()


def list_payroll_payment_history(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    employee_user_id: uuid.UUID | None = None,
) -> list[PayrollPaymentHistoryRow]:
    """Paid payroll history.

    When ``week_start`` is set, filter by that payroll period week only.
    When ``date_from`` / ``date_to`` are set without ``week_start``, filter paid items whose
    payroll period ``week_start`` falls in that inclusive range (not ``paid_at``).
    """
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    _assert_valid_range_filter(
        db_session,
        company_id=company_id,
        employee_user_id=employee_user_id,
    )
    if week_start is not None:
        rows = list_paid_items_for_company_payment_history(
            db_session,
            company_id=company_id,
            payroll_week_start=week_start,
            employee_user_id=employee_user_id,
        )
    else:
        if date_from is None or date_to is None:
            raise PayrollError("date_from and date_to are required when week_start is not provided.")
        if date_from > date_to:
            raise PayrollError("date_from must be before or equal to date_to.")
        rows = list_paid_items_for_company_payment_history(
            db_session,
            company_id=company_id,
            payroll_week_start_from=date_from,
            payroll_week_start_to=date_to,
            employee_user_id=employee_user_id,
        )
    out: list[PayrollPaymentHistoryRow] = []
    for item, period in rows:
        if item.status != "paid" or item.paid_at is None:
            continue
        email, name, _job_title = _employee_display(db_session, item.user_id)
        out.append(
            PayrollPaymentHistoryRow(
                item_id=item.id,
                user_id=item.user_id,
                employee_email=email,
                employee_name=name,
                paid_at=item.paid_at,
                week_start=period.week_start,
                week_end=_week_end_display(period.week_start),
                gross_amount=_decimal_or_none(item.gross_amount),
                cis_tax_amount=_effective_tax_amount_for_item(item),
                net_paid_amount=_effective_net_amount_for_item(item),
                payment_mode=_stored_payment_mode_or_none(item.payment_mode),
                payment_mode_label=_payment_mode_label_for_item(item),
                status=item.status,
                can_open_payslip=True,
                can_undo_paid=True,
            )
        )
    return out


def _assert_valid_range_filter(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    employee_user_id: uuid.UUID | None,
) -> None:
    if employee_user_id is None:
        return
    target = get_user_by_id(db_session, employee_user_id)
    if (
        target is None
        or target.company_id != company_id
        or target.system_role != SystemRole.EMPLOYEE
    ):
        raise PayrollError("Invalid employee filter.")


def _range_shift_rows(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    date_from: date,
    date_to: date,
    employee_user_id: uuid.UUID | None,
) -> tuple[list[dict[str, Any]], int, set[uuid.UUID]]:
    if date_from > date_to:
        raise PayrollError("date_from must be before or equal to date_to.")
    policy = ensure_company_time_policy(db_session, company_id)
    start_utc, end_utc = _date_range_bounds_utc(policy.timezone_name, date_from, date_to)
    rows = list_completed_time_shifts_for_company_range(
        db_session,
        company_id=company_id,
        range_start_utc=start_utc,
        range_end_utc=end_utc,
        user_id=employee_user_id,
    )
    out: list[dict[str, Any]] = []
    total_rounded_seconds = 0
    employee_ids: set[uuid.UUID] = set()
    for shift, location, owner, profile in rows:
        shift_policy = effective_time_policy_for_shift(db_session, shift, location)
        profile_early = bool(profile.early_access_enabled) if profile is not None else False
        early_access = effective_early_access_for_shift(
            db_session,
            location,
            profile_early_access=profile_early,
        )
        metrics = compute_shift_metrics(
            clock_in_at_utc=shift.clock_in_at,
            clock_out_at_utc=shift.clock_out_at,
            break_seconds_tracked=int(shift.break_seconds or 0),
            early_access_enabled=early_access,
            policy=shift_policy,
        )
        rounded = int(metrics.rounded_seconds or 0)
        actual = int(metrics.actual_seconds or 0)
        total_rounded_seconds += rounded
        employee_ids.add(owner.id)
        name = " ".join(
            part
            for part in (
                getattr(profile, "first_name", None),
                getattr(profile, "last_name", None),
            )
            if part
        ).strip()
        out.append(
            {
                "row_type": "shift",
                "employee": name or owner.email or "Employee",
                "employee_email": owner.email or "",
                "role": (getattr(profile, "job_title", None) or "").strip() or "—",
                "period": _payroll_week_start_for_dt(shift.clock_in_at, policy.timezone_name).isoformat(),
                "shift_date": shift.clock_in_at.astimezone(_policy_zone_name(policy.timezone_name)).date().isoformat(),
                "clock_in": shift.clock_in_at.isoformat(),
                "clock_out": shift.clock_out_at.isoformat() if shift.clock_out_at is not None else "",
                "location": getattr(location, "name", "") or "",
                "status": shift.status,
                "hours": f"{rounded / 3600:.2f}",
                "actual_hours": f"{actual / 3600:.2f}",
                "ot_hours": "",
                "gross": "",
                "cis_tax": "",
                "net": "",
                "other_deductions": "",
            },
        )
    return out, total_rounded_seconds, employee_ids


def _range_payroll_total_rows(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    date_from: date,
    date_to: date,
    employee_user_id: uuid.UUID | None,
) -> tuple[list[dict[str, Any]], Decimal | None, Decimal | None, Decimal | None]:
    week_rows: list[dict[str, Any]] = []
    gross_sum = Decimal(0)
    cis_sum = Decimal(0)
    net_sum = Decimal(0)
    has_gross = has_cis = has_net = False
    for week_start in _complete_week_starts_in_range(date_from, date_to):
        period = get_period_by_company_week(db_session, company_id, week_start)
        if period is None:
            continue
        items = list_items_for_period(db_session, period.id)
        if employee_user_id is not None:
            items = [i for i in items if i.user_id == employee_user_id]
        for item in items:
            email, name, job_title = _employee_display(db_session, item.user_id)
            eff_cis = _effective_tax_amount_for_item(item)
            eff_net = _effective_net_amount_for_item(item)
            gross = _decimal_or_none(item.gross_amount)
            if gross is not None:
                gross_sum += gross
                has_gross = True
            if eff_cis is not None:
                cis_sum += eff_cis
                has_cis = True
            if eff_net is not None:
                net_sum += eff_net
                has_net = True
            week_rows.append(
                {
                    "row_type": "payroll_week_total",
                    "employee": (name or email or "Employee"),
                    "employee_email": email or "",
                    "role": job_title or "—",
                    "period": f"{week_start.isoformat()} to {_week_end_display(week_start).isoformat()}",
                    "shift_date": "",
                    "clock_in": "",
                    "clock_out": "",
                    "location": "",
                    "status": item.status,
                    "hours": f"{item.rounded_total_seconds / 3600:.2f}",
                    "actual_hours": "",
                    "ot_hours": f"{item.overtime_seconds / 3600:.2f}",
                    "gross": "" if gross is None else f"{gross:.2f}",
                    "cis_tax": "" if eff_cis is None else f"{eff_cis:.2f}",
                    "net": "" if eff_net is None else f"{eff_net:.2f}",
                    "other_deductions": f"{Decimal(str(item.other_deductions_amount or 0)):.2f}",
                },
            )
    return (
        week_rows,
        gross_sum if has_gross else None,
        cis_sum if has_cis else None,
        net_sum if has_net else None,
    )


def _employee_filter_label(
    db_session: Session,
    *,
    company_id: uuid.UUID,
    employee_user_id: uuid.UUID | None,
) -> str:
    if employee_user_id is None:
        return "All employees"
    target = get_user_by_id(db_session, employee_user_id)
    if target is None or target.company_id != company_id:
        return "Selected employee"
    name = _employee_primary_name(db_session, employee_user_id)
    return f"{name} ({target.email})" if target.email else name


def export_csv_report(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    employee_user_id: uuid.UUID | None = None,
) -> str:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    if date_from is not None or date_to is not None:
        if date_from is None or date_to is None:
            raise PayrollError("date_from and date_to are required together.")
        _assert_valid_range_filter(db_session, company_id=company_id, employee_user_id=employee_user_id)
        company = get_company_by_id(db_session, company_id)
        company_name = company.name if company is not None else "Company"
        shift_rows, total_shift_seconds, employee_ids = _range_shift_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        payroll_rows, gross_total, cis_total, net_total = _range_payroll_total_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        partial_note = PARTIAL_RANGE_PAYROLL_NOTE if _range_has_partial_week_portion(date_from, date_to) else ""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(
            [
                "row_type",
                "company_name",
                "date_from",
                "date_to",
                "employee_filter",
                "employee",
                "employee_email",
                "role",
                "payroll_week_or_period",
                "shift_date",
                "clock_in",
                "clock_out",
                "location",
                "hours",
                "actual_hours",
                "overtime_hours",
                "gross_pay",
                "cis_tax",
                "net_pay",
                "status",
                "notes",
            ],
        )
        employee_label = _employee_filter_label(
            db_session,
            company_id=company_id,
            employee_user_id=employee_user_id,
        )
        for row in shift_rows:
            writer.writerow(
                [
                    row["row_type"],
                    company_name,
                    date_from,
                    date_to,
                    employee_label,
                    row["employee"],
                    row["employee_email"],
                    row["role"],
                    row["period"],
                    row["shift_date"],
                    row["clock_in"],
                    row["clock_out"],
                    row["location"],
                    row["hours"],
                    row["actual_hours"],
                    row["ot_hours"],
                    row["gross"],
                    row["cis_tax"],
                    row["net"],
                    row["status"],
                    partial_note,
                ],
            )
        for row in payroll_rows:
            writer.writerow(
                [
                    row["row_type"],
                    company_name,
                    date_from,
                    date_to,
                    employee_label,
                    row["employee"],
                    row["employee_email"],
                    row["role"],
                    row["period"],
                    row["shift_date"],
                    row["clock_in"],
                    row["clock_out"],
                    row["location"],
                    row["hours"],
                    row["actual_hours"],
                    row["ot_hours"],
                    row["gross"],
                    row["cis_tax"],
                    row["net"],
                    row["status"],
                    "Stored payroll total for a complete payroll week inside the selected range.",
                ],
            )
        if not shift_rows and not payroll_rows:
            writer.writerow(
                [
                    "note",
                    company_name,
                    date_from,
                    date_to,
                    employee_label,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    seconds_to_hours_csv(total_shift_seconds),
                    "",
                    "",
                    gross_total or "",
                    cis_total or "",
                    net_total or "",
                    "",
                    partial_note or "No completed shifts or complete payroll-week totals for this range.",
                ],
            )
        create_internal_audit_event(
            db_session=db_session,
            actor=actor,
            action="payroll.report_exported",
            entity_type="payroll_period",
            entity_id=None,
            company_id=company_id,
            details={
                "export_type": "csv_range",
                "date_from": str(date_from),
                "date_to": str(date_to),
                "row_count": len(shift_rows),
                "employee_user_id": str(employee_user_id) if employee_user_id else None,
            },
        )
        return buffer.getvalue()

    if week_start is None:
        raise PayrollError("week_start is required.")
    report = get_payroll_report(
        db_session,
        actor,
        company_id=company_id,
        week_start=week_start,
        auto_recalculate_if_safe=False,
    )
    company = get_company_by_id(db_session, company_id)
    company_name = company.name if company is not None else "Company"
    week_end = _week_end_display(week_start)
    tz_name = report.period.timezone_name if report.period.total_items else ""

    by_id: dict[uuid.UUID, PayrollItem] = {}
    if report.items:
        pid = report.items[0].period_id
        by_id = {i.id: i for i in list_items_for_period(db_session, pid)}

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        [
            "company_name",
            "week_start",
            "week_end",
            "timezone",
            "employee_email",
            "employee_name",
            "employee_job_title",
            "payment_mode",
            "regular_hours",
            "overtime_hours",
            "total_rounded_hours",
            "hourly_rate_snapshot",
            "tax_rate_snapshot",
            "gross_amount",
            "cis_tax",
            "other_deductions",
            "net_amount",
            "status",
            "rate_missing",
        ],
    )
    for row in report.items:
        item = by_id.get(row.id)
        eff_cis = _effective_tax_amount_for_item(item) if item is not None else None
        eff_net = _effective_net_amount_for_item(item) if item is not None else None
        writer.writerow(
            [
                company_name,
                str(week_start),
                str(week_end),
                tz_name,
                row.employee_email or "",
                row.employee_name or "",
                row.employee_job_title or "",
                _payment_mode_label(row.payment_mode),
                seconds_to_hours_csv(row.regular_seconds),
                seconds_to_hours_csv(row.overtime_seconds),
                seconds_to_hours_csv(row.rounded_total_seconds),
                row.hourly_rate_snapshot,
                row.tax_rate_snapshot,
                row.gross_amount,
                eff_cis,
                row.other_deductions_amount,
                eff_net,
                row.status,
                row.rate_missing,
            ],
        )
    period_entity = str(report.period.id) if report.period.total_items else None
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.report_exported",
        entity_type="payroll_period",
        entity_id=period_entity,
        company_id=company_id,
        details={
            "export_type": "csv",
            "week_start": str(week_start),
            "row_count": len(report.items),
        },
    )
    return buffer.getvalue()


def _decimal_from_export_value(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "—":
        return None
    try:
        return Decimal(text.replace("£", "").replace(",", ""))
    except Exception:
        return None


def _float_from_export_value(value: object) -> float | None:
    dec = _decimal_from_export_value(value)
    return float(dec) if dec is not None else None


def _date_from_export_value(value: object) -> date | str:
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return date.fromisoformat(text)
    except ValueError:
        return text


def _xlsx_status_label(status: object) -> str:
    raw = str(status or "").strip()
    return raw.replace("_", " ").title() if raw else ""


def _append_payroll_xlsx_sheet(
    *,
    workbook,
    company_name: str,
    period_label: str,
    timezone_name: str,
    employee_filter_label: str,
    generated_at: datetime,
    summary: dict[str, object],
    rows: list[dict[str, object]],
) -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = workbook.active
    ws.title = "Payroll Report"

    title_fill = PatternFill("solid", fgColor="111827")
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    label_fill = PatternFill("solid", fgColor="F3F4F6")
    white_font = Font(color="FFFFFF", bold=True, size=14)
    bold = Font(bold=True)

    ws.merge_cells("A1:U1")
    ws["A1"] = "TimIQ Payroll Report"
    ws["A1"].font = white_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    meta_rows = [
        ("Company", company_name),
        ("Period", period_label),
        ("Employee filter", employee_filter_label),
        ("Generated", generated_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("Timezone", timezone_name or "—"),
    ]
    row_idx = 3
    for label, value in meta_rows:
        ws.cell(row=row_idx, column=1, value=label).font = bold
        ws.cell(row=row_idx, column=1).fill = label_fill
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Summary").font = Font(bold=True, size=12)
    row_idx += 1
    for label, value, number_format in (
        ("Total hours", summary.get("total_hours"), "0.00"),
        ("Employees", summary.get("employee_count"), "0"),
        ("Gross pay", summary.get("gross_pay"), '£#,##0.00;[Red]-£#,##0.00'),
        ("CIS tax", summary.get("cis_tax"), '£#,##0.00;[Red]-£#,##0.00'),
        ("Net pay", summary.get("net_pay"), '£#,##0.00;[Red]-£#,##0.00'),
    ):
        ws.cell(row=row_idx, column=1, value=label).font = bold
        ws.cell(row=row_idx, column=1).fill = label_fill
        cell = ws.cell(row=row_idx, column=2, value=value)
        cell.number_format = number_format
        row_idx += 1

    row_idx += 2
    header_row = row_idx
    headers = [
        "Row type",
        "Company",
        "Date from",
        "Date to",
        "Employee filter",
        "Employee",
        "Employee email",
        "Role",
        "Payroll week / period",
        "Shift date",
        "Clock in",
        "Clock out",
        "Location",
        "Hours",
        "Actual hours",
        "Overtime hours",
        "Gross pay",
        "CIS tax",
        "Net pay",
        "Status",
        "Notes",
    ]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    data_start = header_row + 1
    if rows:
        for out_row, row in enumerate(rows, start=data_start):
            values = [
                row.get("row_type"),
                company_name,
                row.get("date_from"),
                row.get("date_to"),
                employee_filter_label,
                row.get("employee"),
                row.get("employee_email"),
                row.get("role"),
                row.get("period"),
                _date_from_export_value(row.get("shift_date")),
                row.get("clock_in"),
                row.get("clock_out"),
                row.get("location"),
                _float_from_export_value(row.get("hours")),
                _float_from_export_value(row.get("actual_hours")),
                _float_from_export_value(row.get("ot_hours")),
                _float_from_export_value(row.get("gross")),
                _float_from_export_value(row.get("cis_tax")),
                _float_from_export_value(row.get("net")),
                _xlsx_status_label(row.get("status")),
                row.get("notes"),
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row=out_row, column=col, value=value)
    else:
        ws.cell(row=data_start, column=1, value="No payable payroll rows for this selected range.")
        ws.merge_cells(start_row=data_start, start_column=1, end_row=data_start, end_column=len(headers))

    date_cols = (3, 4, 10)
    hours_cols = (14, 15, 16)
    money_cols = (17, 18, 19)
    for row in ws.iter_rows(min_row=data_start, max_row=max(data_start, data_start + max(len(rows), 1) - 1)):
        for cell in row:
            if cell.column in date_cols:
                cell.number_format = "yyyy-mm-dd"
            elif cell.column in hours_cols:
                cell.number_format = "0.00"
            elif cell.column in money_cols:
                cell.number_format = '£#,##0.00;[Red]-£#,##0.00'
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = [18, 24, 12, 12, 24, 26, 30, 18, 24, 12, 22, 22, 24, 10, 12, 12, 13, 13, 13, 14, 42]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:U{max(data_start, data_start + max(len(rows), 1) - 1)}"
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_xlsx_report(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    employee_user_id: uuid.UUID | None = None,
) -> bytes:
    from openpyxl import Workbook

    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    company = get_company_by_id(db_session, company_id)
    company_name = company.name if company is not None else "Company"
    generated_at = datetime.now(timezone.utc)

    if date_from is not None or date_to is not None:
        if date_from is None or date_to is None:
            raise PayrollError("date_from and date_to are required together.")
        if date_from > date_to:
            raise PayrollError("date_from must be before or equal to date_to.")
        _assert_valid_range_filter(db_session, company_id=company_id, employee_user_id=employee_user_id)
        policy = ensure_company_time_policy(db_session, company_id)
        shift_rows, total_shift_seconds, employee_ids = _range_shift_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        payroll_rows, gross_total, cis_total, net_total = _range_payroll_total_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        employee_label = _employee_filter_label(
            db_session,
            company_id=company_id,
            employee_user_id=employee_user_id,
        )
        note = PARTIAL_RANGE_PAYROLL_NOTE if _range_has_partial_week_portion(date_from, date_to) else ""
        rows = [
            {
                **row,
                "date_from": date_from,
                "date_to": date_to,
                "notes": note,
            }
            for row in shift_rows
        ]
        rows.extend(
            {
                **row,
                "date_from": date_from,
                "date_to": date_to,
                "notes": "Stored payroll total for a complete payroll week inside the selected range.",
            }
            for row in payroll_rows
        )
        summary = {
            "total_hours": total_shift_seconds / 3600,
            "employee_count": len(employee_ids),
            "gross_pay": float(gross_total) if gross_total is not None else None,
            "cis_tax": float(cis_total) if cis_total is not None else None,
            "net_pay": float(net_total) if net_total is not None else None,
        }
        body = _append_payroll_xlsx_sheet(
            workbook=Workbook(),
            company_name=company_name,
            period_label=f"{date_from.isoformat()} to {date_to.isoformat()}",
            timezone_name=policy.timezone_name,
            employee_filter_label=employee_label,
            generated_at=generated_at,
            summary=summary,
            rows=rows,
        )
        create_internal_audit_event(
            db_session=db_session,
            actor=actor,
            action="payroll.report_exported",
            entity_type="payroll_period",
            entity_id=None,
            company_id=company_id,
            details={
                "export_type": "xlsx_range",
                "date_from": str(date_from),
                "date_to": str(date_to),
                "row_count": len(rows),
                "employee_user_id": str(employee_user_id) if employee_user_id else None,
            },
        )
        return body

    if week_start is None:
        raise PayrollError("week_start is required.")
    report = get_payroll_report(
        db_session,
        actor,
        company_id=company_id,
        week_start=week_start,
        auto_recalculate_if_safe=False,
    )
    week_end = _week_end_display(week_start)
    by_id: dict[uuid.UUID, PayrollItem] = {}
    if report.items:
        pid = report.items[0].period_id
        by_id = {i.id: i for i in list_items_for_period(db_session, pid)}
    rows: list[dict[str, object]] = []
    gross_sum = cis_sum = net_sum = Decimal(0)
    has_gross = has_cis = has_net = False
    for row in report.items:
        if employee_user_id is not None and row.user_id != employee_user_id:
            continue
        item = by_id.get(row.id)
        eff_cis = _effective_tax_amount_for_item(item) if item is not None else None
        eff_net = _effective_net_amount_for_item(item) if item is not None else None
        if row.gross_amount is not None:
            gross_sum += Decimal(str(row.gross_amount))
            has_gross = True
        if eff_cis is not None:
            cis_sum += eff_cis
            has_cis = True
        if eff_net is not None:
            net_sum += eff_net
            has_net = True
        rows.append(
            {
                "row_type": "payroll_week_total",
                "date_from": week_start,
                "date_to": week_end,
                "employee": row.employee_name or row.employee_email or "Employee",
                "employee_email": row.employee_email or "",
                "role": row.employee_job_title or "—",
                "period": f"{week_start.isoformat()} to {week_end.isoformat()}",
                "shift_date": "",
                "clock_in": "",
                "clock_out": "",
                "location": "",
                "hours": row.rounded_total_seconds / 3600,
                "actual_hours": "",
                "ot_hours": row.overtime_seconds / 3600,
                "gross": row.gross_amount,
                "cis_tax": eff_cis,
                "net": eff_net,
                "status": row.status,
                "notes": "",
            },
        )
    employee_label = _employee_filter_label(db_session, company_id=company_id, employee_user_id=employee_user_id)
    summary = {
        "total_hours": sum(float(_float_from_export_value(r.get("hours")) or 0) for r in rows),
        "employee_count": len({r.get("employee_email") for r in rows if r.get("employee_email")}),
        "gross_pay": float(gross_sum) if has_gross else None,
        "cis_tax": float(cis_sum) if has_cis else None,
        "net_pay": float(net_sum) if has_net else None,
    }
    body = _append_payroll_xlsx_sheet(
        workbook=Workbook(),
        company_name=company_name,
        period_label=f"{week_start.isoformat()} to {week_end.isoformat()}",
        timezone_name=report.period.timezone_name,
        employee_filter_label=employee_label,
        generated_at=generated_at,
        summary=summary,
        rows=rows,
    )
    period_entity = str(report.period.id) if report.period.total_items else None
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.report_exported",
        entity_type="payroll_period",
        entity_id=period_entity,
        company_id=company_id,
        details={
            "export_type": "xlsx",
            "week_start": str(week_start),
            "row_count": len(rows),
            "employee_user_id": str(employee_user_id) if employee_user_id else None,
        },
    )
    return body


def _payroll_report_alert_lines(alerts: PayrollReportAlerts) -> list[str]:
    lines: list[str] = []
    if alerts.payroll_period_not_calculated:
        lines.append("Payroll not calculated for this week yet.")
    if alerts.payroll_needs_recalculation:
        lines.append("Time records changed; payroll may need recalculation.")
    if alerts.pending_approval_count:
        lines.append(f"Pending approval: {alerts.pending_approval_count}")
    if alerts.rate_missing_employees_count:
        lines.append(f"Employees with missing rate: {alerts.rate_missing_employees_count}")
    if alerts.open_shifts_started_in_week_count:
        lines.append(f"Open shifts started in week: {alerts.open_shifts_started_in_week_count}")
    if alerts.zero_rounded_hours_employees_count:
        lines.append(f"Employees with zero rounded hours: {alerts.zero_rounded_hours_employees_count}")
    return lines


def export_print_html(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date,
    user_id: uuid.UUID | None = None,
) -> str:
    company = get_company_by_id(db_session, company_id)
    name = html.escape(company.name if company else "Company")
    report = get_payroll_report(
        db_session,
        actor,
        company_id=company_id,
        week_start=week_start,
        user_id=user_id,
        auto_recalculate_if_safe=False,
    )
    week_end = _week_end_display(week_start)
    week_end_esc = html.escape(str(week_end))
    wk_esc = html.escape(str(week_start))
    tz_esc = html.escape(report.period.timezone_name if report.period.total_items else "")

    by_id: dict[uuid.UUID, PayrollItem] = {}
    if report.items:
        pid = report.items[0].period_id
        by_id = {i.id: i for i in list_items_for_period(db_session, pid)}

    rows_html: list[str] = []
    for row in report.items:
        item = by_id.get(row.id)
        eff_cis = _effective_tax_amount_for_item(item) if item is not None else None
        eff_net = _effective_net_amount_for_item(item) if item is not None else None
        cis_txt = "—" if eff_cis is None else f"{eff_cis:.2f}"
        net_txt = "—" if eff_net is None else f"{eff_net:.2f}"
        mode_lbl = html.escape(_payment_mode_label(row.payment_mode))
        jt = html.escape((row.employee_job_title or "").strip())
        jt_cell = jt if jt else "—"
        rows_html.append(
            "<tr>"
            f"<td>{html.escape(row.employee_email or '')}</td>"
            f"<td>{html.escape(row.employee_name or '')}</td>"
            f"<td>{jt_cell}</td>"
            f"<td>{mode_lbl}</td>"
            f"<td>{row.regular_seconds / 3600:.2f}</td>"
            f"<td>{row.overtime_seconds / 3600:.2f}</td>"
            f"<td>{row.rounded_total_seconds / 3600:.2f}</td>"
            f"<td>{row.gross_amount if row.gross_amount is not None else '—'}</td>"
            f"<td>{cis_txt}</td>"
            f"<td>{html.escape(str(row.other_deductions_amount))}</td>"
            f"<td>{net_txt}</td>"
            f"<td>{html.escape(row.status)}</td>"
            "</tr>",
        )
    html_out = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"/><title>Payroll {name} — {wk_esc}</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 24px; color: #111; }}
h1 {{ font-size: 1.25rem; }}
table {{ border-collapse: collapse; width: 100%; margin-top: 16px; }}
th, td {{ border: 1px solid #ccc; padding: 8px; text-align: left; font-size: 0.875rem; }}
th {{ background: #f4f4f5; }}
@media print {{ body {{ margin: 12px; }} }}
</style></head><body>
<h1>Payroll — {name}</h1>
<p>Week {wk_esc} to {week_end_esc} · {tz_esc}</p>
<table><thead><tr>
<th>Email</th><th>Name</th><th>Job title</th><th>Payment mode</th>
<th>Regular h</th><th>OT h</th><th>Rounded h</th><th>Gross</th><th>CIS tax</th><th>Other ded.</th><th>Net</th><th>Status</th>
</tr></thead><tbody>
{"".join(rows_html)}
</tbody></table>
<p style="margin-top:16px;font-size:12px;color:#666;">Use browser Print → Save as PDF for a PDF copy.</p>
</body></html>"""
    period_entity = str(report.period.id) if report.period.total_items else None
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.report_exported",
        entity_type="payroll_period",
        entity_id=period_entity,
        company_id=company_id,
        details={
            "export_type": "print_html",
            "week_start": str(week_start),
            "row_count": len(report.items),
            "user_id": str(user_id) if user_id else None,
        },
    )
    return html_out


def export_pdf_report(
    db_session: Session,
    actor: User,
    *,
    company_id: uuid.UUID,
    week_start: date | None = None,
    user_id: uuid.UUID | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    employee_user_id: uuid.UUID | None = None,
) -> bytes:
    assert_payroll_admin_or_administrator(actor)
    assert_payroll_company_scope(actor, company_id)
    company = get_company_by_id(db_session, company_id)
    company_name = company.name if company is not None else "Company"
    if date_from is not None or date_to is not None:
        if date_from is None or date_to is None:
            raise PayrollError("date_from and date_to are required together.")
        _assert_valid_range_filter(db_session, company_id=company_id, employee_user_id=employee_user_id)
        policy = ensure_company_time_policy(db_session, company_id)
        shift_rows, total_shift_seconds, employee_ids = _range_shift_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        payroll_rows, gross_total, cis_total, net_total = _range_payroll_total_rows(
            db_session,
            company_id=company_id,
            date_from=date_from,
            date_to=date_to,
            employee_user_id=employee_user_id,
        )
        alert_lines: list[str] = []
        if _range_has_partial_week_portion(date_from, date_to):
            alert_lines.append(PARTIAL_RANGE_PAYROLL_NOTE)
        if payroll_rows:
            alert_lines.append("Stored pay totals are shown only for complete payroll weeks fully inside this range.")
        pdf_rows = [
            {
                "employee": row["employee"],
                "role": row["role"],
                "period": row["shift_date"],
                "hours": row["hours"],
                "ot_hours": row["ot_hours"] or "—",
                "gross": "—",
                "cis_tax": "—",
                "net": "—",
                "other_deductions": "—",
                "status": row["status"],
            }
            for row in shift_rows
        ]
        pdf_rows.extend(
            {
                "employee": row["employee"],
                "role": row["role"],
                "period": row["period"],
                "hours": row["hours"],
                "ot_hours": row["ot_hours"],
                "gross": row["gross"] or "—",
                "cis_tax": row["cis_tax"] or "—",
                "net": row["net"] or "—",
                "other_deductions": row["other_deductions"] or "—",
                "status": row["status"],
            }
            for row in payroll_rows
        )
        body = build_payroll_report_pdf(
            company_name=company_name,
            week_start=date_from,
            week_end=date_to,
            timezone_name=policy.timezone_name,
            rows=pdf_rows,
            total_hours_seconds=total_shift_seconds,
            total_gross=gross_total,
            total_cis_tax=cis_total,
            total_net=net_total,
            alert_lines=alert_lines,
            period_label=f"Date range: {date_from.isoformat()} to {date_to.isoformat()} · {policy.timezone_name}",
            employee_filter_label=_employee_filter_label(
                db_session,
                company_id=company_id,
                employee_user_id=employee_user_id,
            ),
            employee_count=len(employee_ids),
        )
        create_internal_audit_event(
            db_session=db_session,
            actor=actor,
            action="payroll.report_exported",
            entity_type="payroll_period",
            entity_id=None,
            company_id=company_id,
            details={
                "export_type": "pdf_range",
                "date_from": str(date_from),
                "date_to": str(date_to),
                "row_count": len(shift_rows),
                "employee_user_id": str(employee_user_id) if employee_user_id else None,
            },
        )
        return body

    if week_start is None:
        raise PayrollError("week_start is required.")
    report = get_payroll_report(
        db_session,
        actor,
        company_id=company_id,
        week_start=week_start,
        user_id=user_id,
        auto_recalculate_if_safe=False,
    )
    week_end = _week_end_display(week_start)
    tz_name = report.period.timezone_name if report.period.total_items else ""

    by_id: dict[uuid.UUID, PayrollItem] = {}
    if report.items:
        pid = report.items[0].period_id
        by_id = {i.id: i for i in list_items_for_period(db_session, pid)}

    pdf_rows: list[dict[str, Any]] = []
    total_seconds = 0
    gross_sum = Decimal(0)
    cis_sum = Decimal(0)
    net_sum = Decimal(0)
    has_gross = has_cis = has_net = False
    for row in report.items:
        item = by_id.get(row.id)
        eff_cis = _effective_tax_amount_for_item(item) if item is not None else None
        eff_net = _effective_net_amount_for_item(item) if item is not None else None
        emp_label = (row.employee_name or row.employee_email or "").strip() or "—"
        jt = (row.employee_job_title or "").strip() or "—"
        total_seconds += row.rounded_total_seconds
        if row.gross_amount is not None:
            gross_sum += Decimal(str(row.gross_amount))
            has_gross = True
        if eff_cis is not None:
            cis_sum += eff_cis
            has_cis = True
        if eff_net is not None:
            net_sum += eff_net
            has_net = True
        pdf_rows.append(
            {
                "employee": emp_label,
                "role": jt,
                "hours": f"{row.rounded_total_seconds / 3600:.2f}",
                "ot_hours": f"{row.overtime_seconds / 3600:.2f}",
                "gross": "—" if row.gross_amount is None else f"{row.gross_amount:.2f}",
                "cis_tax": "—" if eff_cis is None else f"{eff_cis:.2f}",
                "net": "—" if eff_net is None else f"{eff_net:.2f}",
                "other_deductions": f"{row.other_deductions_amount:.2f}",
                "status": row.status,
            },
        )

    body = build_payroll_report_pdf(
        company_name=company_name,
        week_start=week_start,
        week_end=week_end,
        timezone_name=tz_name,
        rows=pdf_rows,
        total_hours_seconds=total_seconds,
        total_gross=gross_sum if has_gross else None,
        total_cis_tax=cis_sum if has_cis else None,
        total_net=net_sum if has_net else None,
        alert_lines=_payroll_report_alert_lines(report.alerts),
    )
    period_entity = str(report.period.id) if report.period.total_items else None
    create_internal_audit_event(
        db_session=db_session,
        actor=actor,
        action="payroll.report_exported",
        entity_type="payroll_period",
        entity_id=period_entity,
        company_id=company_id,
        details={
            "export_type": "pdf",
            "week_start": str(week_start),
            "row_count": len(report.items),
            "user_id": str(user_id) if user_id else None,
        },
    )
    return body
